from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.colors import HexColor, black
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
import asyncio
import json  # ✅ qo'shib qo'ying (fayl boshida ham bo'lishi mumkin)
import logging
import sys
import time
import re
import os
import tempfile
from datetime import datetime, timezone, timedelta, date, time as dtime
from collections import defaultdict, deque

import aiosqlite
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from aiogram import Bot, Dispatcher, Router, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import CommandStart, StateFilter
from aiogram.types import (
    Message, CallbackQuery,
    ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton,
    FSInputFile
)
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.dispatcher.middlewares.base import BaseMiddleware


# ===================== CONFIG =====================
BOT_TOKEN = "8699119508:AAHqgX_HPd4Uf7y0BGjWwjknONz63Gus5ME"  # ⚠️ BotFather -> /revoke -> yangi token qo'ying
ADMIN_ID = 6551039574
ACADEMY_NAME = "Smart Edu Center"
CERT_LOGO_PATH = "assets/logo.png"       # loyihangizda shu fayl bo‘lsin
CERT_SIGN_PATH = "assets/sign.png"       # imzo rasmi (PNG)
CERT_MIN_PERCENT = 60.0                  # oddiy testlarda threshold
# ===================== PAYMENT CONFIG =====================
PAYMENT_CARD = "9860 0121 0681 0950"
PAYMENT_OWNER = "Jasurbek Aktamov"
PAYMENT_PHONE = "+998 93 224 47 30"
PAYMENT_NOTE = "To'lov izohiga test kodini yozing (masalan: MS-001)."
DB_NAME = "edutest.db"

TOP_N = 10

# Asia/Tashkent fixed offset
TZ = timezone(timedelta(hours=5))  # +05:00

# Anti-spam / security
MSG_WINDOW_SEC = 15
MSG_LIMIT = 10
CB_WINDOW_SEC = 15
CB_LIMIT = 15
SAME_CALLBACK_COOLDOWN_SEC = 2
MUTE_STEPS = [120, 300, 900, 3600, 86400]  # 2m, 5m, 15m, 1h, 1d
STRIKE_RESET_AFTER_SEC = 24 * 3600

# Countdown (reverse timer)
COUNTDOWN_TICK_SEC = 60  # 60s


# ===================== UI TEXT =====================
BTN_FREE = "🆓 Tekin testlar"
BTN_PAID = "🧾 Pullik testlar"
BTN_MY = "📦 Mening testlarim"
BTN_RATING = "🏆 Reyting"
BTN_RESULTS = "📄 Natijalarim"
BTN_HELP = "ℹ️ Yordam"

BTN_ADMIN_ADD = "➕ Test Qo'shish (Admin)"
BTN_ADMIN_PENDING = "🟡 Pending to'lovlar (Admin)"
BTN_ADMIN_STATS = "📊 Statistika (Admin)"
BTN_ADMIN_CHANNELS = "📣 Majburiy kanallar (Admin)"  # ✅ NEW

BTN_HOME = "🏠 Bosh menyu"
BTN_BACK = "⬅️ Ortga"
BTN_CANCEL = "Bekor qilish"

TOP_MENU_BTNS = {
    BTN_FREE, BTN_PAID, BTN_MY, BTN_RATING, BTN_RESULTS, BTN_HELP,
    BTN_ADMIN_ADD, BTN_ADMIN_PENDING, BTN_ADMIN_STATS, BTN_ADMIN_CHANNELS,
    BTN_HOME, BTN_BACK, BTN_CANCEL
}


# ===================== STATES =====================
class RegistrationState(StatesGroup):
    full_name = State()
    phone = State()


class PaymentState(StatesGroup):
    waiting_for_screenshot = State()
    test_id = State()


class AdminAddTestState(StatesGroup):
    subject = State()  # ✅ NEW
    title = State()
    code = State()
    is_free = State()
    price = State()
    duration = State()
    questions_count = State()
    exam_type = State()
    schedule_mode = State()
    schedule_date = State()
    schedule_time = State()
    file = State()
    answers = State()


class TestProcessState(StatesGroup):
    solving = State()


class AdminRejectState(StatesGroup):
    waiting_reason = State()
    payment_id = State()


class SearchByCodeState(StatesGroup):
    waiting_code = State()
    scope = State()  # paid | free | my


# ✅ NEW: Admin channels management
class AdminChannelState(StatesGroup):
    waiting_channel_ref = State()   # @username yoki -100...
    waiting_join_url = State()      # optional
    edit_id = State()               # which row to edit
    waiting_edit_payload = State()  # 2 lines: title + link


# ===================== HELPERS =====================
def now_str_local():
    return datetime.now(TZ).strftime("%Y-%m-%d %H:%M")


def now_ts() -> int:
    return int(time.time())


def fmt_seconds(seconds: int) -> str:
    seconds = max(0, int(seconds))
    if seconds < 60:
        return f"{seconds} soniya"
    m, s = divmod(seconds, 60)
    if m < 60:
        return f"{m} daqiqa {s} soniya"
    h, m = divmod(m, 60)
    if h < 24:
        return f"{h} soat {m} daqiqa"
    d, h = divmod(h, 24)
    return f"{d} kun {h} soat"


def normalize_code(code: str) -> str:
    code = (code or "").strip().upper()
    code = re.sub(r"\s+", "", code)
    return code


def is_valid_code(code: str) -> bool:
    return bool(re.fullmatch(r"[A-Z0-9\-_]{3,20}", code))


def normalize_answers(ans: str) -> str:
    ans = (ans or "").strip().lower()
    ans = re.sub(r"\s+", "", ans)
    return ans


def is_answer_string_valid(ans: str) -> bool:
    return bool(re.fullmatch(r"[a-z]+", ans))


def parse_date(date_str: str):
    try:
        return datetime.strptime(date_str.strip(), "%Y-%m-%d").date()
    except:
        return None


def parse_time(time_str: str) -> dtime | None:
    try:
        return datetime.strptime(time_str.strip(), "%H:%M").time()
    except:
        return None


def to_start_ts(date_str: str, time_str: str) -> tuple[int, str] | None:
    """
    Returns: (start_ts_utc, start_at_string_local) or None if invalid
    """
    d = parse_date(date_str)
    t = parse_time(time_str)
    if not d or not t:
        return None
    dt_local = datetime(d.year, d.month, d.day, t.hour, t.minute, tzinfo=TZ)
    start_at = dt_local.strftime("%Y-%m-%d %H:%M")
    return int(dt_local.timestamp()), start_at
def is_certificate_eligible(exam_type: str, percent: float, grade: str | None) -> bool:
    et = (exam_type or "simple")
    if et in ("rasch", "maxsus"):
        return grade is not None
    return float(percent or 0.0) >= float(CERT_MIN_PERCENT)

async def get_user_full_name(user_id_db: int) -> str:
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT full_name FROM users WHERE id=?", (user_id_db,)) as cur:
            row = await cur.fetchone()
    return (row[0] if row and row[0] else "Noma'lum")

async def get_test_meta_for_certificate(test_id: int):
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT code, title, subject, start_mode, start_ts, duration, exam_type
            FROM tests WHERE id=?
        """, (test_id,)) as cur:
            return await cur.fetchone()

async def is_admin_user(telegram_id: int) -> bool:
    if telegram_id == ADMIN_ID:
        return True
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT role FROM users WHERE telegram_id=?", (telegram_id,)) as cur:
            row = await cur.fetchone()
            return bool(row and row[0] == "admin")


async def get_user_db_id(telegram_id: int):
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT id FROM users WHERE telegram_id=?", (telegram_id,)) as cur:
            row = await cur.fetchone()
            return row[0] if row else None


async def test_by_code(code: str):
    code = normalize_code(code)
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT id, code, title, is_free, price, duration, questions_count, file_id, answers,
                   start_mode, start_ts, start_at
            FROM tests
            WHERE code=?
        """, (code,)) as cur:
            return await cur.fetchone()


async def test_code_exists(code: str) -> bool:
    code = normalize_code(code)
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT 1 FROM tests WHERE code=? LIMIT 1", (code,)) as cur:
            row = await cur.fetchone()
            return bool(row)


def is_scheduled_not_started(start_mode: str, start_ts_val: int | None) -> tuple[bool, int]:
    """
    Returns (blocked, seconds_left)
    """
    if (start_mode or "normal") != "scheduled":
        return (False, 0)
    if not start_ts_val:
        return (False, 0)
    diff = int(start_ts_val) - now_ts()
    if diff > 0:
        return (True, diff)
    return (False, 0)

def is_scheduled_expired(start_mode: str, start_ts_val: int | None, duration_min: int | None) -> tuple[bool, int]:
    """
    Returns (expired, seconds_after_end)
    Faqat scheduled test uchun ishlaydi.
    """
    if (start_mode or "normal") != "scheduled":
        return (False, 0)
    if not start_ts_val:
        return (False, 0)
    dur = int(duration_min or 0) * 60
    if dur <= 0:
        return (False, 0)

    end_ts = int(start_ts_val) + dur
    diff = now_ts() - end_ts
    if diff > 60:  # 60 sekund grace
        return (True, diff)
    return (False, 0)

def make_certificate_pdf(
    full_name: str,
    subject: str,
    test_code: str,
    test_title: str,
    exam_date_str: str,
    percent: float = 0.0,
    correct_count: int = 0,
    total_questions: int = 0,
    rasch_ball: float | None = None,
    grade: str | None = None,
    out_path: str = "certificate.pdf"
):
    """
    CHIROYLI ALBOMNIY (LANDSCAPE) SERTIFIKAT
    """
    # Landscape A4 (842 × 595 pt)
    page = landscape(A4)
    c = canvas.Canvas(out_path, pagesize=page)
    w, h = page

    # ================== ORQA FON (agar bersangiz) ==================
    # try:
    #     bg = ImageReader("assets/cert_background.jpg")   # o'zingizning nomingiz
    #     c.drawImage(bg, 0, 0, width=w, height=h, preserveAspectRatio=True, mask="auto")
    # except:
    #     pass

    # ================== RAMKALAR ==================
    # Tashqi oltin ramka
    c.setStrokeColor(HexColor("#D4AF37"))  # oltin
    c.setLineWidth(10)
    c.rect(20, 20, w - 40, h - 40)

    # Ichki ko‘k ramka
    c.setStrokeColor(HexColor("#1a2a6c"))
    c.setLineWidth(4)
    c.rect(45, 45, w - 90, h - 90)

    # ================== LOGO ==================
    try:
        if CERT_LOGO_PATH and os.path.exists(CERT_LOGO_PATH):
            logo = ImageReader(CERT_LOGO_PATH)
            c.drawImage(logo, 65, h - 145, width=120, height=120, mask="auto")
    except:
        pass

    # ================== SARLAVHA ==================
    c.setFont("Helvetica-Bold", 55)
    c.setFillColor(HexColor("#1a2a6c"))
    c.drawCentredString(w/2, h - 105, "SERTIFIKAT")

    c.setFont("Helvetica", 19)
    c.setFillColor(HexColor("#555555"))
    c.drawCentredString(w/2, h - 145, f"{ACADEMY_NAME} tomonidan taqdim etiladi")

    # ================== ISMI ==================
    c.setFont("Helvetica-Bold", 32)
    c.setFillColor(black)
    c.drawCentredString(w/2, h - 205, full_name.upper())

    c.setFont("Helvetica", 15)
    c.drawCentredString(w/2, h - 240, "ga Ushbu sertifikat quyidagi yutug'i uchun berildi")

    # ================== BATAFSIL MA'LUMOTLAR ==================
    y = h - 290
    lh = 32  # line height

    c.setFont("Helvetica-Bold", 17)
    c.drawString(95, y, "Fan:")
    c.setFont("Helvetica", 17)
    c.drawString(247, y, subject)

    y -= lh
    c.setFont("Helvetica-Bold", 17)
    c.drawString(95, y, "Test nomi:")
    c.setFont("Helvetica", 17)
    c.drawString(247, y, test_title)

    y -= lh
    c.setFont("Helvetica-Bold", 17)
    c.drawString(95, y, "Test kodi:")
    c.setFont("Helvetica", 17)
    c.drawString(247, y, test_code)

    y -= lh
    c.setFont("Helvetica-Bold", 17)
    c.drawString(95, y, "Topshirilgan sana: ")
    c.setFont("Helvetica", 17)
    c.drawString(247, y, exam_date_str)

    # ================== NATIJA BLOKI ==================
    y = h - 430
    c.setFont("Helvetica-Bold", 22)
    c.setFillColor(HexColor("#006400"))  # yashil
    c.drawCentredString(w/2, y, f"NATIJA: {correct_count}/{total_questions}  —  {percent:.1f}%")

    if grade:
        c.setFont("Helvetica-Bold", 38)
        c.setFillColor(HexColor("#D4AF37"))
        c.drawCentredString(w/2, y - 65, f"BAHO: {grade}")

    if rasch_ball is not None:
        c.setFont("Helvetica", 16)
        c.setFillColor(black)
        c.drawCentredString(w/2, y - 105, f"Rasch ball: {rasch_ball:.1f}")

    # ================== IMZO VA MUHR ==================
    sign_y = 108

    # Chap tomonda imzo
    c.setFont("Helvetica-Bold", 15)
    c.drawString(95, sign_y - 48, "Direktor: Po'latov J.")

    try:
        if CERT_SIGN_PATH and os.path.exists(CERT_SIGN_PATH):
            sign = ImageReader(CERT_SIGN_PATH)
            c.drawImage(sign, w - 280, sign_y - 38, width=190, height=75, mask="auto")
    except:
        pass



    # Pastki yozuv
    c.setFont("Helvetica", 10)
    c.setFillColor(HexColor("#666666"))
    c.drawCentredString(w/2, 32, "© 2026 Smart Edu Center • Raqamli imzo bilan tasdiqlangan")

    c.showPage()
    c.save()

# ===================== RUSH MODULE (Finalize after exam end) =====================


# === NEW: Rasch stabilizatsiya sozlamalari ===
RUSH_SMALL_N = 20        # <=20 bo'lsa "kichik guruh" rejimi
RUSH_LARGE_N = 50        # >=50 bo'lsa "katta guruh" rejimi (hozirgi percentile ishlaydi)
# === NEW: Juda kichik N uchun to'liq fallback ===
RUSH_HARD_FALLBACK_N = 6   # <= 6 qatnashuvchi bo'lsa Rasch emas, oddiy foizdan ball


# p smoothing (Bayes) parametrlari: p = (correct + A) / (total + A + B)
RUSH_SMOOTH_A = 1.0
RUSH_SMOOTH_B = 1.0

# kichik guruhda theta->ball rescale uchun barqaror diapazon
RUSH_FIXED_LO = -3.0
RUSH_FIXED_HI =  3.0

# b_list ni haddan tashqari ketib qolmasligi uchun clamp
B_MIN = -4.0
B_MAX =  4.0


# ===================== RASCH (RUSH) MODULE - FIXED =====================

# Milliy sertifikat Rasch ball chegaralari (BBA/UZBMB e'loniga mos)
# 70+ A+, 65-69.9 A, 60-64.9 B+, 55-59.9 B, 50-54.9 C+, 46-49.9 C
# 46 dan past -> sertifikat/daraja yo'q
def grade_from_rasch_ball(ball: float) -> str | None:
    ball = float(ball or 0.0)
    if ball >= 70:
        return "A+"
    if ball >= 65:
        return "A"
    if ball >= 60:
        return "B+"
    if ball >= 55:
        return "B"
    if ball >= 50:
        return "C+"
    if ball >= 46:
        return "C"
    return None  # <= 45.9 -> umuman daraja yo'q (sertifikat bermasligi mumkin)

import math

THETA_MIN = -6.0
THETA_MAX = 6.0

def _sigmoid(x: float) -> float:
    if x >= 35:
        return 1.0
    if x <= -35:
        return 0.0
    return 1.0 / (1.0 + math.exp(-x))

def _clamp(x: float, a: float, b: float) -> float:
    return a if x < a else b if x > b else x

def rasch_difficulty_from_p(p: float) -> float:
    p = _clamp(float(p), 0.01, 0.99)
    return math.log((1.0 - p) / p)
def rasch_difficulty_from_counts(correct: int, total: int, a: float = RUSH_SMOOTH_A, b: float = RUSH_SMOOTH_B) -> float:
    """
    Kichik N uchun p ni smoothing qilamiz:
    p = (correct + a) / (total + a + b)
    """
    total = int(total or 0)
    correct = int(correct or 0)
    if total <= 0:
        return 0.0
    p = (correct + float(a)) / (total + float(a) + float(b))
    bi = rasch_difficulty_from_p(p)
    return _clamp(bi, B_MIN, B_MAX)


def choose_rescale_bounds(thetas: list[float]) -> tuple[float, float]:
    """
    N kichik bo'lsa fixed lo/hi.
    N katta bo'lsa hozirgi 5%-95% percentile.
    Oraliqda bo'lsa mean±2*std (barqarorroq).
    """
    n = len(thetas)
    if n <= 0:
        return (RUSH_FIXED_LO, RUSH_FIXED_HI)

    thetas_sorted = sorted(thetas)

    def _pct(arr, q):
        idx = int(round((len(arr) - 1) * q))
        idx = max(0, min(len(arr) - 1, idx))
        return float(arr[idx])

    if n <= RUSH_SMALL_N:
        return (RUSH_FIXED_LO, RUSH_FIXED_HI)

    if n >= RUSH_LARGE_N:
        lo = _pct(thetas_sorted, 0.05)
        hi = _pct(thetas_sorted, 0.95)
        if hi <= lo + 1e-9:
            return (RUSH_FIXED_LO, RUSH_FIXED_HI)
        return (lo, hi)

    # Oraliq N: mean ± 2*std
    mean = sum(thetas) / n
    var = sum((x - mean) ** 2 for x in thetas) / max(1, n - 1)
    std = math.sqrt(var)
    lo = mean - 2.0 * std
    hi = mean + 2.0 * std
    lo = _clamp(lo, THETA_MIN, THETA_MAX)
    hi = _clamp(hi, THETA_MIN, THETA_MAX)
    if hi <= lo + 1e-9:
        return (RUSH_FIXED_LO, RUSH_FIXED_HI)
    return (lo, hi)


def estimate_theta_rasch(ans: str, key: str, b_list: list[float], iters: int = 25) -> float:
    theta = 0.0
    n = len(b_list)

    for _ in range(iters):
        g = 0.0
        h = 0.0
        for i in range(n):
            x = 1.0 if ans[i] == key[i] else 0.0
            p = _sigmoid(theta - b_list[i])
            g += (x - p)
            h -= (p * (1.0 - p))

        if abs(h) < 1e-9:
            break

        step = g / h
        theta -= step
        theta = _clamp(theta, THETA_MIN, THETA_MAX)

        if abs(step) < 1e-4:
            break

    return theta

def rescale_theta_to_75(theta: float, lo: float, hi: float) -> float:
    if hi <= lo + 1e-9:
        return 37.5
    x = (theta - lo) / (hi - lo)
    x = _clamp(x, 0.0, 1.0)
    return x * 75.0


def rasch_ball_from_weighted_ratio(ratio_0_1: float) -> float:
    """
    Siz hisoblayotgan weighted ratio (num/den) 0..1 bo'ladi.
    Rasch ballni 0..75 shkalaga o'tkazib beramiz.
    """
    ratio_0_1 = max(0.0, min(1.0, float(ratio_0_1 or 0.0)))
    return ratio_0_1 * 75.0
def fallback_rasch_ball_from_raw_percent(raw_percent_0_100: float) -> float:
    """
    Juda kichik N bo'lsa Raschni ishlatmaymiz.
    Oddiy foizni 0..75 Rasch ballga o'tkazamiz (chiziqli).
    """
    p = max(0.0, min(100.0, float(raw_percent_0_100 or 0.0)))
    return (p / 100.0) * 75.0




async def get_test_end_ts(test_id: int) -> int:
    """
    Testning umumiy tugash vaqtini hisoblaydi:
    - scheduled bo'lsa: start_ts + duration*60
    - normal bo'lsa: earliest finished_ts + duration*60 (fallback)
    """
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute(
            "SELECT start_mode, start_ts, duration FROM tests WHERE id=?",
            (test_id,)
        ) as cur:
            t = await cur.fetchone()

        if not t:
            return 0

        start_mode, start_ts_val, duration_min = t
        duration_sec = int(duration_min or 0) * 60

        if (start_mode or "normal") == "scheduled" and int(start_ts_val or 0) > 0:
            return int(start_ts_val) + duration_sec

        # fallback: 1chi topshirgan vaqt + duration
        async with db.execute("""
            SELECT MIN(finished_ts)
            FROM results
            WHERE test_id=? AND finished_ts>0
        """, (test_id,)) as cur:
            r = await cur.fetchone()
        first_finish = int((r[0] or 0) if r else 0)
        if first_finish <= 0:
            return 0
        return first_finish + duration_sec


async def finalize_rush_for_test_if_ready(test_id: int) -> bool:
    end_ts = await get_test_end_ts(test_id)
    if end_ts <= 0 or now_ts() < end_ts:
        return False

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT rush_finalized FROM tests WHERE id=?", (test_id,)) as cur:
            row = await cur.fetchone()
        if row and int(row[0] or 0) == 1:
            return True

        async with db.execute("SELECT questions_count, answers FROM tests WHERE id=?", (test_id,)) as cur:
            t = await cur.fetchone()
        if not t:
            return False

        qcount = int(t[0] or 0)
        correct_key = normalize_answers(t[1] or "")
        qcount = min(qcount, len(correct_key))
        if qcount <= 0:
            return False

        async with db.execute("""
            SELECT id, user_answers
            FROM results
            WHERE test_id=? AND user_answers IS NOT NULL AND LENGTH(user_answers) >= ?
        """, (test_id, qcount)) as cur:
            rows = await cur.fetchall()

        if not rows:
            return False

        solvers = []
        for rid, ans_raw in rows:
            ans = normalize_answers(ans_raw or "")
            if len(ans) >= qcount:
                solvers.append(ans)

        n_solvers = len(solvers)

        # === HARD FALLBACK ===
        if n_solvers <= RUSH_HARD_FALLBACK_N:
            for rid, ans_raw in rows:
                ans = normalize_answers(ans_raw or "")
                if len(ans) < qcount:
                    continue

                score = 0
                for i in range(qcount):
                    if ans[i] == correct_key[i]:
                        score += 1

                raw_percent = (score / qcount) * 100.0
                rasch_ball = fallback_rasch_ball_from_raw_percent(raw_percent)
                rasch_percent = (rasch_ball / 75.0) * 100.0
                grade = grade_from_rasch_ball(rasch_ball)

                await db.execute("""
                    UPDATE results
                    SET rasch_ball=?, rasch_percent=?, grade=?
                    WHERE id=?
                """, (float(rasch_ball), float(rasch_percent), grade, int(rid)))

            # ✅ finalize 1 marta
            await db.execute(
                "UPDATE tests SET rush_finalized=1, rush_finalized_ts=? WHERE id=?",
                (now_ts(), test_id)
            )
            await db.commit()
            return True

        # --- item difficulty ---
        correct_counts = [0] * qcount
        total_counts = [0] * qcount

        for ans in solvers:
            for i in range(qcount):
                total_counts[i] += 1
                if ans[i] == correct_key[i]:
                    correct_counts[i] += 1

        b_list = []
        for i in range(qcount):
            if total_counts[i] == 0:
                b_list.append(0.0)
                continue
            if n_solvers <= RUSH_SMALL_N:
                b_list.append(rasch_difficulty_from_counts(correct_counts[i], total_counts[i]))
            else:
                p = correct_counts[i] / total_counts[i]
                b_list.append(_clamp(rasch_difficulty_from_p(p), B_MIN, B_MAX))

        # --- bounds ---
        thetas = [estimate_theta_rasch(ans, correct_key, b_list) for ans in solvers]
        lo, hi = choose_rescale_bounds(thetas)

        # --- update ALL results ---
        for rid, ans_raw in rows:
            ans = normalize_answers(ans_raw or "")
            if len(ans) < qcount:
                continue

            theta = estimate_theta_rasch(ans, correct_key, b_list)
            rasch_ball = rescale_theta_to_75(theta, lo, hi)
            rasch_percent = (rasch_ball / 75.0) * 100.0
            grade = grade_from_rasch_ball(rasch_ball)

            await db.execute("""
                UPDATE results
                SET rasch_ball=?, rasch_percent=?, grade=?
                WHERE id=?
            """, (float(rasch_ball), float(rasch_percent), grade, int(rid)))

        # ✅ finalize 1 marta (LOOPDAN KEYIN!)
        await db.execute(
            "UPDATE tests SET rush_finalized=1, rush_finalized_ts=? WHERE id=?",
            (now_ts(), test_id)
        )
        await db.commit()
        return True



# ===================== MAXSUS MODULE (Rasch-like, but local) =====================

MAXSUS_RELIABILITY_M = 3 # n kichik bo'lsa natijani 50% ga yaqinlashtiradi
def _round_half_up(x: float) -> int:
    # 10.5 -> 11 (python round() banker's rounding emas)
    return int(math.floor(x + 0.5))

def maxsus_grade_from_k_any(n: int, k: int) -> str | None:
    n = int(n or 0)
    k = int(k or 0)
    if n <= 0:
        return None

    # 1) ANIQ jadval: 30 / 35 / 55 (rasmdagi kabi)
    if n == 30:
        if 25 <= k <= 30: return "A+"
        if 22 <= k <= 24: return "A"
        if 18 <= k <= 21: return "B+"
        if 15 <= k <= 17: return "B"
        if 12 <= k <= 14: return "C+"
        if 9  <= k <= 11: return "C"
        return None

    if n == 35:
        if 29 <= k <= 35: return "A+"
        if 25 <= k <= 28: return "A"
        if 21 <= k <= 24: return "B+"
        if 18 <= k <= 20: return "B"
        if 14 <= k <= 17: return "C+"
        if 11 <= k <= 13: return "C"
        return None

    if n == 55:
        if 45 <= k <= 55: return "A+"
        if 40 <= k <= 44: return "A"
        if 34 <= k <= 39: return "B+"
        if 28 <= k <= 33: return "B"
        if 22 <= k <= 27: return "C+"
        if 17 <= k <= 21: return "C"
        return None

    # 2) UNIVERSAL (boshqa n lar uchun)
    c_min   = _round_half_up(0.30 * n)
    cp_min  = _round_half_up(0.40 * n)
    b_min   = _round_half_up(0.50 * n)
    bp_min  = _round_half_up((0.62 if n >= 50 else 0.60) * n)
    a_min   = _round_half_up(0.72 * n)
    ap_min  = _round_half_up(0.82 * n)

    # monoton bo'lsin
    cp_min = max(cp_min, c_min + 1)
    b_min  = max(b_min,  cp_min + 1)
    bp_min = max(bp_min, b_min + 1)
    a_min  = max(a_min,  bp_min + 1)
    ap_min = max(ap_min, a_min + 1)

    if k >= ap_min: return "A+"
    if k >= a_min:  return "A"
    if k >= bp_min: return "B+"
    if k >= b_min:  return "B"
    if k >= cp_min: return "C+"
    if k >= c_min:  return "C"
    return None

def maxsus_theta_from_score(k: int, n: int) -> float:
    k = int(k)
    n = int(n)
    if n <= 0:
        return 0.0
    return math.log((k + 0.5) / ((n - k) + 0.5))

def maxsus_rasch_p(theta: float) -> float:
    return _sigmoid(theta)

def maxsus_reliability(n: int, m: int = MAXSUS_RELIABILITY_M) -> float:
    n = int(n)
    m = int(m)
    if n <= 0:
        return 0.0
    return n / (n + m)

def maxsus_ball_percent_from_score(k: int, n: int) -> tuple[float, float, float]:
    theta = maxsus_theta_from_score(k, n)
    p = maxsus_rasch_p(theta)  # 0..1
    r = maxsus_reliability(n)  # 0..1
    p_final = 0.5 + r * (p - 0.5)  # 0..1
    p_final = max(0.0, min(1.0, p_final))

    rasch_ball = 75.0 * p_final  # 0..75
    rasch_percent = 100.0 * p_final  # 0..100
    return rasch_ball, rasch_percent, theta

async def finalize_maxsus_for_test_if_ready(test_id: int) -> bool:
    end_ts = await get_test_end_ts(test_id)
    if end_ts <= 0:
        return False
    if now_ts() < end_ts:
        return False

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT maxsus_finalized FROM tests WHERE id=?", (test_id,)) as cur:
            row = await cur.fetchone()
        if row and int(row[0] or 0) == 1:
            return True

        async with db.execute("SELECT questions_count, answers FROM tests WHERE id=?", (test_id,)) as cur:
            t = await cur.fetchone()
        if not t:
            return False

        qcount = int(t[0] or 0)
        correct_key = normalize_answers(t[1] or "")
        qcount = min(qcount, len(correct_key))
        if qcount <= 0:
            return False

        async with db.execute("""
            SELECT id, user_answers
            FROM results
            WHERE test_id=? AND user_answers IS NOT NULL AND LENGTH(user_answers) >= ?
        """, (test_id, qcount)) as cur:
            rows = await cur.fetchall()

        if not rows:
            return False

        for rid, ans_raw in rows:
            ans = normalize_answers(ans_raw or "")
            if len(ans) < qcount:
                continue

            k = 0
            for i in range(qcount):
                if ans[i] == correct_key[i]:
                    k += 1

            rasch_ball, rasch_percent, _theta = maxsus_ball_percent_from_score(k, qcount)
            grade = maxsus_grade_from_k_any(qcount, k)  # ✅ ASOSIY O'ZGARISH (jadval bo'yicha)

            await db.execute("""
                UPDATE results
                SET rasch_ball=?, rasch_percent=?, grade=?
                WHERE id=?
            """, (float(rasch_ball), float(rasch_percent), grade, int(rid)))

        await db.execute(
            "UPDATE tests SET maxsus_finalized=1, maxsus_finalized_ts=? WHERE id=?",
            (now_ts(), test_id)
        )
        await db.commit()

    return True

# ===================== REQUIRED CHANNELS (FORCED SUBSCRIPTION) =====================
def _normalize_channel_ref(text: str) -> str:
    s = (text or "").strip()
    if s.startswith("@"):
        s = s[1:]
    return s


def _looks_like_int(s: str) -> bool:
    try:
        int(s)
        return True
    except:
        return False


async def get_required_channels(active_only: bool = True):
    q = "SELECT id, chat_id, username, title, join_url, is_active FROM required_channels"
    if active_only:
        q += " WHERE is_active=1"
    q += " ORDER BY id ASC"
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute(q) as cur:
            return await cur.fetchall()


def build_join_kb(channels_rows):
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for (_id, chat_id, username, title, join_url, _is_active) in channels_rows:
        url = (join_url or "").strip()
        if not url:
            if username:
                url = f"https://t.me/{username}"
            else:
                url = ""
        btn_text = f"➕ {title or (('@'+username) if username else str(chat_id))}"
        if url:
            kb.inline_keyboard.append([InlineKeyboardButton(text=btn_text, url=url)])
    kb.inline_keyboard.append([InlineKeyboardButton(text="✅ A’zo bo‘ldim — Tekshirish", callback_data="chk_sub")])
    return kb


async def check_user_subscribed(bot: Bot, user_id: int) -> tuple[bool, list]:
    rows = await get_required_channels(active_only=True)
    if not rows:
        return True, []

    not_joined = []
    for row in rows:
        (_id, chat_id, username, title, join_url, is_active) = row

        # If chat_id is missing but username exists, try to resolve and cache it.
        resolved_chat_id = chat_id
        if not resolved_chat_id and username:
            try:
                ch = await bot.get_chat(f"@{username}")
                resolved_chat_id = ch.id
                async with aiosqlite.connect(DB_NAME) as db:
                    await db.execute("UPDATE required_channels SET chat_id=? WHERE id=?", (int(resolved_chat_id), _id))
                    await db.commit()
            except Exception:
                resolved_chat_id = None

        if not resolved_chat_id:
            # Can't verify membership without a chat_id (private groups/channels require bot to be in the chat).
            not_joined.append(row)
            continue

        try:
            member = await bot.get_chat_member(chat_id=int(resolved_chat_id), user_id=user_id)
            status = getattr(member, "status", None)
            if status in ("left", "kicked"):
                not_joined.append(row)
        except Exception:
            not_joined.append(row)

    return (len(not_joined) == 0), not_joined


async def ensure_subscribed_message(message: Message, bot: Bot) -> bool:
    if await is_admin_user(message.from_user.id):
        return True
    ok, not_joined = await check_user_subscribed(bot, message.from_user.id)
    if ok:
        return True
    kb = build_join_kb(not_joined)
    await message.answer(
        "🔒 <b>Botdan foydalanish uchun quyidagi kanal/guruhlarga a’zo bo‘ling:</b>\n\n"
        "A’zo bo‘lgach, <b>“Tekshirish”</b> tugmasini bosing.",
        parse_mode=ParseMode.HTML,
        reply_markup=kb
    )
    return False


async def ensure_subscribed_callback(callback: CallbackQuery, bot: Bot) -> bool:
    if await is_admin_user(callback.from_user.id):
        return True
    ok, not_joined = await check_user_subscribed(bot, callback.from_user.id)
    if ok:
        return True
    kb = build_join_kb(not_joined)
    await callback.message.answer(
        "🔒 <b>Davom etish uchun majburiy kanal/guruhlarga a’zo bo‘ling:</b>\n\n"
        "A’zo bo‘lgach, <b>“Tekshirish”</b> tugmasini bosing.",
        parse_mode=ParseMode.HTML,
        reply_markup=kb
    )
    return False


# ✅ NEW: Global middlewares (enforce for both channels and groups)
class SubscriptionMessageMiddleware(BaseMiddleware):
    async def __call__(self, handler, event: Message, data):
        if not event.from_user:
            return await handler(event, data)

        bot: Bot = data.get("bot")
        if not bot:
            return await handler(event, data)

        # Admin bypass
        if await is_admin_user(event.from_user.id):
            return await handler(event, data)

        ok = await ensure_subscribed_message(event, bot)
        if not ok:
            return
        return await handler(event, data)


class SubscriptionCallbackMiddleware(BaseMiddleware):
    async def __call__(self, handler, event: CallbackQuery, data):
        if not event.from_user:
            return await handler(event, data)

        # Always allow the re-check button.
        if (event.data or "") == "chk_sub":
            return await handler(event, data)

        bot: Bot = data.get("bot")
        if not bot:
            return await handler(event, data)

        # Admin bypass
        if await is_admin_user(event.from_user.id):
            return await handler(event, data)

        ok = await ensure_subscribed_callback(event, bot)
        if not ok:
            return
        return await handler(event, data)


# ===================== DATABASE INIT + MIGRATIONS =====================
async def _ensure_column(db: aiosqlite.Connection, table: str, column: str, col_def: str):
    async with db.execute(f"PRAGMA table_info({table})") as cur:
        cols = await cur.fetchall()
    col_names = {c[1] for c in cols}
    if column not in col_names:
        await db.execute(f"ALTER TABLE {table} ADD COLUMN {column} {col_def}")


async def init_db():
    async with aiosqlite.connect(DB_NAME) as db:
        await db.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                telegram_id INTEGER UNIQUE,
                full_name TEXT,
                phone TEXT,
                role TEXT DEFAULT 'student'
            )
        """)

        await db.execute("""
            CREATE TABLE IF NOT EXISTS tests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE,
                title TEXT,
                is_free INTEGER DEFAULT 0,
                price INTEGER DEFAULT 0,
                duration INTEGER DEFAULT 0,
                questions_count INTEGER DEFAULT 30,
                file_id TEXT,
                answers TEXT,

                start_mode TEXT DEFAULT 'normal',
                start_ts INTEGER DEFAULT 0,
                start_at TEXT,

                rush_finalized INTEGER DEFAULT 0,
                rush_finalized_ts INTEGER DEFAULT 0
            )
        """)

        await db.execute("""
            CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                test_id INTEGER,
                status TEXT DEFAULT 'pending',
                screenshot_id TEXT,
                created_at TEXT DEFAULT (datetime('now')),
                reject_reason TEXT,
                started_at TEXT,
                started_ts INTEGER,
                FOREIGN KEY(user_id) REFERENCES users(id),
                FOREIGN KEY(test_id) REFERENCES tests(id)
            )
        """)

        await db.execute("""
            CREATE TABLE IF NOT EXISTS results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                test_id INTEGER,
                score INTEGER,
                total_questions INTEGER,
                percent REAL,
                date TEXT,
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
        """)

        await db.execute("""
            CREATE TABLE IF NOT EXISTS sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                test_id INTEGER,
                mode TEXT, -- free/paid
                start_ts INTEGER,
                duration_sec INTEGER,
                created_at TEXT,
                UNIQUE(user_id, test_id, mode)
            )
        """)

        await db.execute("""
            CREATE TABLE IF NOT EXISTS user_limits (
                user_tg_id INTEGER PRIMARY KEY,
                strikes INTEGER DEFAULT 0,
                blocked_until_ts INTEGER DEFAULT 0,
                last_violation_ts INTEGER DEFAULT 0
            )
        """)

        # ✅ NEW: required channels
        await db.execute("""
            CREATE TABLE IF NOT EXISTS required_channels (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                chat_id INTEGER UNIQUE,
                username TEXT,
                title TEXT,
                join_url TEXT,
                is_active INTEGER DEFAULT 1
            )
        """)

        # ✅ RUSH: tests finalize flags
        await _ensure_column(db, "tests", "rush_finalized", "INTEGER DEFAULT 0")
        await _ensure_column(db, "tests", "rush_finalized_ts", "INTEGER DEFAULT 0")
        # ✅ MAXSUS: finalize flags
        await _ensure_column(db, "tests", "maxsus_finalized", "INTEGER DEFAULT 0")
        await _ensure_column(db, "tests", "maxsus_finalized_ts", "INTEGER DEFAULT 0")
        # ✅ EXAM TYPES
        await _ensure_column(db, "tests", "exam_type", "TEXT DEFAULT 'simple'")
        await _ensure_column(db, "tests", "dtm_cfg", "TEXT")

        # ✅ POINTS BASED RESULTS
        await _ensure_column(db, "results", "earned_points", "REAL")
        await _ensure_column(db, "results", "max_points", "REAL")
        await _ensure_column(db, "results", "details", "TEXT")

        # ✅ RUSH: results table migrations
        await _ensure_column(db, "results", "user_answers", "TEXT")

        await _ensure_column(db, "results", "grade", "TEXT")
        await _ensure_column(db, "results", "finished_ts", "INTEGER DEFAULT 0")

        # ✅ RUSH: tests finalize flags
        await _ensure_column(db, "tests", "rush_finalized", "INTEGER DEFAULT 0")
        await _ensure_column(db, "tests", "rush_finalized_ts", "INTEGER DEFAULT 0")

        # migrations for tests
        await _ensure_column(db, "tests", "is_free", "INTEGER DEFAULT 0")
        await _ensure_column(db, "tests", "code", "TEXT")
        await _ensure_column(db, "tests", "questions_count", "INTEGER DEFAULT 30")

        # ✅ RASCH: alohida ustunlar
        await _ensure_column(db, "results", "rasch_ball", "REAL")
        await _ensure_column(db, "results", "rasch_percent", "REAL")


        # scheduled fields
        await _ensure_column(db, "tests", "start_mode", "TEXT DEFAULT 'normal'")
        await _ensure_column(db, "tests", "start_ts", "INTEGER DEFAULT 0")
        await _ensure_column(db, "tests", "start_at", "TEXT")
        # ✅ SUBJECT (FAN)
        await _ensure_column(db, "tests", "subject", "TEXT")

        # ✅ CERTIFICATE FIELDS
        await _ensure_column(db, "results", "certificate_sent", "INTEGER DEFAULT 0")
        await _ensure_column(db, "results", "certificate_file_id", "TEXT")
        await _ensure_column(db, "results", "certificate_sent_ts", "INTEGER DEFAULT 0")

        await db.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_tests_code_unique ON tests(code)")

        await db.commit()


# ===================== SECURITY =====================
class SecurityManager:
    def __init__(self):
        self.msg_hits: dict[int, deque] = defaultdict(deque)
        self.cb_hits: dict[int, deque] = defaultdict(deque)
        self.last_cb: dict[int, tuple[str, float]] = {}

    async def _get_limit_row(self, user_id: int):
        async with aiosqlite.connect(DB_NAME) as db:
            async with db.execute(
                "SELECT strikes, blocked_until_ts, last_violation_ts FROM user_limits WHERE user_tg_id=?",
                (user_id,),
            ) as cur:
                row = await cur.fetchone()
            if not row:
                await db.execute(
                    "INSERT INTO user_limits (user_tg_id, strikes, blocked_until_ts, last_violation_ts) VALUES (?, 0, 0, 0)",
                    (user_id,),
                )
                await db.commit()
                return (0, 0, 0)
            return row

    async def is_blocked(self, user_id: int) -> tuple[bool, int]:
        _, blocked_until_ts, _ = await self._get_limit_row(user_id)
        now_ = int(time.time())
        if blocked_until_ts and blocked_until_ts > now_:
            return True, blocked_until_ts - now_
        return False, 0

    async def _apply_violation(self, user_id: int):
        now_ = int(time.time())
        strikes, _, last_violation_ts = await self._get_limit_row(user_id)

        if last_violation_ts and (now_ - last_violation_ts) > STRIKE_RESET_AFTER_SEC:
            strikes = 0

        strikes += 1
        step_index = min(strikes - 1, len(MUTE_STEPS) - 1)
        mute_sec = MUTE_STEPS[step_index]
        new_blocked_until = now_ + mute_sec

        async with aiosqlite.connect(DB_NAME) as db:
            await db.execute(
                "UPDATE user_limits SET strikes=?, blocked_until_ts=?, last_violation_ts=? WHERE user_tg_id=?",
                (strikes, new_blocked_until, now_, user_id),
            )
            await db.commit()

        return strikes, mute_sec

    def _trim_deque(self, dq: deque, window_sec: int, now_ts_: float):
        while dq and (now_ts_ - dq[0]) > window_sec:
            dq.popleft()

    async def check_message(self, user_id: int) -> tuple[bool, str]:
        blocked, remain = await self.is_blocked(user_id)
        if blocked:
            return False, f"⛔️ Siz vaqtincha bloklandingiz.\n⏳ Qolgan vaqt: {fmt_seconds(remain)}"

        now_ts_ = time.time()
        dq = self.msg_hits[user_id]
        dq.append(now_ts_)
        self._trim_deque(dq, MSG_WINDOW_SEC, now_ts_)

        if len(dq) > MSG_LIMIT:
            strikes, mute_sec = await self._apply_violation(user_id)
            return False, (
                "🚫 Juda ko‘p so‘rov yuboryapsiz.\n"
                f"🔒 Vaqtincha blok: {fmt_seconds(mute_sec)}\n"
                f"⚠️ Ogohlantirishlar: {strikes}"
            )
        return True, ""

    async def check_callback(self, user_id: int, data: str) -> tuple[bool, str, bool]:
        blocked, remain = await self.is_blocked(user_id)
        if blocked:
            return False, f"⛔️ Siz vaqtincha bloklandingiz.\n⏳ Qolgan vaqt: {fmt_seconds(remain)}", False

        now_ts_ = time.time()
        last = self.last_cb.get(user_id)
        if last and last[0] == data and (now_ts_ - last[1]) < SAME_CALLBACK_COOLDOWN_SEC:
            return False, "⏳ Sekinroq bosing (2 soniya).", True

        self.last_cb[user_id] = (data, now_ts_)

        dq = self.cb_hits[user_id]
        dq.append(now_ts_)
        self._trim_deque(dq, CB_WINDOW_SEC, now_ts_)

        if len(dq) > CB_LIMIT:
            strikes, mute_sec = await self._apply_violation(user_id)
            return False, (
                "🚫 Tugmalarni juda tez bosyapsiz.\n"
                f"🔒 Vaqtincha blok: {fmt_seconds(mute_sec)}\n"
                f"⚠️ Ogohlantirishlar: {strikes}"
            ), False

        return True, "", False


SECURITY = SecurityManager()


class MessageSecurityMiddleware(BaseMiddleware):
    async def __call__(self, handler, event: Message, data):
        user_id = event.from_user.id if event.from_user else 0
        if not user_id:
            return await handler(event, data)
        allowed, msg = await SECURITY.check_message(user_id)
        if not allowed:
            try:
                await event.answer(msg)
            except:
                pass
            return
        return await handler(event, data)


class CallbackSecurityMiddleware(BaseMiddleware):
    async def __call__(self, handler, event: CallbackQuery, data):
        user_id = event.from_user.id if event.from_user else 0
        if not user_id:
            return await handler(event, data)
        allowed, msg, silent = await SECURITY.check_callback(user_id, event.data or "")
        if not allowed:
            try:
                await event.answer(msg, show_alert=not silent)
            except:
                pass
            return
        return await handler(event, data)


# ===================== KEYBOARDS =====================
def main_menu_kb(is_admin=False):
    buttons = [
        [KeyboardButton(text=BTN_FREE), KeyboardButton(text=BTN_PAID)],
        [KeyboardButton(text=BTN_MY), KeyboardButton(text=BTN_RATING)],
        [KeyboardButton(text=BTN_RESULTS), KeyboardButton(text=BTN_HELP)],
    ]
    if is_admin:
        buttons.append([KeyboardButton(text=BTN_ADMIN_ADD)])
        buttons.append([KeyboardButton(text=BTN_ADMIN_PENDING)])
        buttons.append([KeyboardButton(text=BTN_ADMIN_STATS)])
        buttons.append([KeyboardButton(text=BTN_ADMIN_CHANNELS)])  # ✅ NEW
    return ReplyKeyboardMarkup(keyboard=buttons, resize_keyboard=True)


def phone_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📞 Kontaktni ulashish", request_contact=True)],
            [KeyboardButton(text=BTN_CANCEL), KeyboardButton(text=BTN_HOME)],
        ],
        resize_keyboard=True
    )


def nav_kb(include_back=True):
    row = [KeyboardButton(text=BTN_HOME)]
    if include_back:
        row.insert(0, KeyboardButton(text=BTN_BACK))
    return ReplyKeyboardMarkup(keyboard=[row, [KeyboardButton(text=BTN_CANCEL)]], resize_keyboard=True)


def search_kb(scope: str):
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔎 Kod orqali topish", callback_data=f"askcode_{scope}")]
    ])


# ===================== COUNTDOWN =====================
ACTIVE_TIMERS: dict[tuple[int, int], asyncio.Task] = {}


async def start_countdown(bot: Bot, chat_id: int, test_id: int, end_ts: int, msg_id: int):
    key = (chat_id, test_id)
    try:
        while True:
            remain = end_ts - now_ts()
            if remain <= 0:
                try:
                    await bot.edit_message_text(
                        chat_id=chat_id,
                        message_id=msg_id,
                        text="⏱ <b>Vaqt tugadi!</b>\nJavob yuborilmasa qabul qilinmaydi.",
                        parse_mode=ParseMode.HTML
                    )
                except:
                    pass
                break

            text = f"⏳ <b>Qolgan vaqt:</b> {fmt_seconds(remain)}"
            try:
                await bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=msg_id,
                    text=text,
                    parse_mode=ParseMode.HTML
                )
            except:
                pass

            await asyncio.sleep(COUNTDOWN_TICK_SEC)
    finally:
        ACTIVE_TIMERS.pop(key, None)


def cancel_timer(user_tg_id: int, test_id: int):
    key = (user_tg_id, test_id)
    task = ACTIVE_TIMERS.pop(key, None)
    if task and not task.done():
        task.cancel()


async def upsert_session(user_id: int, test_id: int, mode: str, start_ts_: int, duration_sec: int):
    async with aiosqlite.connect(DB_NAME) as db:
        await db.execute("""
            INSERT INTO sessions (user_id, test_id, mode, start_ts, duration_sec, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(user_id, test_id, mode) DO UPDATE SET
                start_ts=excluded.start_ts,
                duration_sec=excluded.duration_sec,
                created_at=excluded.created_at
        """, (user_id, test_id, mode, start_ts_, duration_sec, now_str_local()))
        await db.commit()


async def get_session(user_id: int, test_id: int, mode: str):
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT start_ts, duration_sec
            FROM sessions
            WHERE user_id=? AND test_id=? AND mode=?
        """, (user_id, test_id, mode)) as cur:
            return await cur.fetchone()


async def delete_session(user_id: int, test_id: int, mode: str):
    async with aiosqlite.connect(DB_NAME) as db:
        await db.execute("DELETE FROM sessions WHERE user_id=? AND test_id=? AND mode=?", (user_id, test_id, mode))
        await db.commit()


# ===================== ROUTER + MIDDLEWARE =====================
router = Router()
router.message.middleware(MessageSecurityMiddleware())
router.callback_query.middleware(CallbackSecurityMiddleware())

# ✅ Forced subscription as GLOBAL gate
router.message.middleware(SubscriptionMessageMiddleware())
router.callback_query.middleware(SubscriptionCallbackMiddleware())


# ===================== REQUIRED SUB CHECK BUTTON =====================
@router.callback_query(F.data == "chk_sub")
async def check_subscribed_cb(callback: CallbackQuery, bot: Bot, state: FSMContext):
    ok, not_joined = await check_user_subscribed(bot, callback.from_user.id)
    if not ok:
        kb = build_join_kb(not_joined)
        await callback.answer("Hali hamma kanalga a’zo emassiz.", show_alert=True)
        await callback.message.answer("⚠️ Hali a’zo bo‘lmagan kanallar bor:", reply_markup=kb)
        return

    await callback.answer("✅ Hammasi joyida!", show_alert=False)
    await state.clear()
    admin_flag = await is_admin_user(callback.from_user.id)
    await callback.message.answer("✅ Kirish ruxsat berildi.", reply_markup=main_menu_kb(admin_flag))


# ===================== GLOBAL NAVIGATION (STATE RESET) =====================
async def go_home(message: Message, state: FSMContext):
    await state.clear()
    admin_flag = await is_admin_user(message.from_user.id)
    await message.answer("🏠 Bosh menyu", reply_markup=main_menu_kb(admin_flag))


@router.message(StateFilter("*"), F.text == BTN_HOME)
async def nav_home(message: Message, state: FSMContext):
    await go_home(message, state)


@router.message(StateFilter("*"), F.text == BTN_BACK)
async def nav_back(message: Message, state: FSMContext):
    await go_home(message, state)


@router.message(StateFilter("*"), F.text.in_({BTN_FREE, BTN_PAID, BTN_MY, BTN_RATING, BTN_RESULTS, BTN_HELP, BTN_ADMIN_ADD, BTN_ADMIN_PENDING, BTN_ADMIN_STATS, BTN_ADMIN_CHANNELS}))
async def nav_any_section(message: Message, state: FSMContext, bot: Bot):
    # ✅ forced subscription gate (admin bypass is inside)
    if not await ensure_subscribed_message(message, bot):
        return

    await state.clear()

    if message.text == BTN_FREE:
        return await show_free_tests(message)
    if message.text == BTN_PAID:
        return await show_paid_tests(message)
    if message.text == BTN_MY:
        return await my_tests(message)
    if message.text == BTN_RATING:
        return await rating_menu(message)
    if message.text == BTN_RESULTS:
        return await my_results(message)
    if message.text == BTN_HELP:
        return await help_menu(message)
    if message.text == BTN_ADMIN_ADD:
        return await admin_add_test(message, state)
    if message.text == BTN_ADMIN_PENDING:
        return await admin_pending(message)
    if message.text == BTN_ADMIN_STATS:
        return await admin_stats(message)
    if message.text == BTN_ADMIN_CHANNELS:
        return await admin_channels_panel(message, state)


@router.message(StateFilter("*"), F.text == BTN_CANCEL)
async def cancel_action(message: Message, state: FSMContext):
    await state.clear()
    admin_flag = await is_admin_user(message.from_user.id)
    await message.answer("Bekor qilindi.", reply_markup=main_menu_kb(admin_flag))


# ===================== /start + REGISTRATION =====================
@router.message(CommandStart())
async def command_start(message: Message, state: FSMContext, bot: Bot):
    # ✅ forced subscription gate
    if not await ensure_subscribed_message(message, bot):
        return

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT * FROM users WHERE telegram_id=?", (message.from_user.id,)) as cur:
            user = await cur.fetchone()

    if user:
        admin_flag = await is_admin_user(message.from_user.id)
        await message.answer(f"Xush kelibsiz, {user[2]}!", reply_markup=main_menu_kb(admin_flag))
        return

    await message.answer(
        "Assalomu alaykum! Botdan foydalanish uchun ro'yxatdan o'ting.\nIsm-familiyangizni kiriting:",
        reply_markup=nav_kb(include_back=False)
    )
    await state.set_state(RegistrationState.full_name)


@router.message(RegistrationState.full_name)
async def process_name(message: Message, state: FSMContext):
    if (message.text or "") in TOP_MENU_BTNS:
        return
    await state.update_data(full_name=message.text.strip())
    await message.answer("Telefon raqamingizni yuboring:", reply_markup=phone_kb())
    await state.set_state(RegistrationState.phone)


@router.message(RegistrationState.phone)
async def process_phone(message: Message, state: FSMContext):
    contact = message.contact.phone_number if message.contact else (message.text or "").strip()
    data = await state.get_data()
    role = 'admin' if message.from_user.id == ADMIN_ID else 'student'

    async with aiosqlite.connect(DB_NAME) as db:
        await db.execute(
            "INSERT OR IGNORE INTO users (telegram_id, full_name, phone, role) VALUES (?, ?, ?, ?)",
            (message.from_user.id, data.get("full_name", ""), contact, role)
        )
        await db.commit()

    admin_flag = await is_admin_user(message.from_user.id)
    await message.answer("✅ Ro'yxatdan o'tish yakunlandi!", reply_markup=main_menu_kb(admin_flag))
    await state.clear()


# ===================== SEARCH BY CODE =====================
@router.callback_query(F.data.startswith("askcode_"))
async def ask_code(callback: CallbackQuery, state: FSMContext, bot: Bot):
    if not await ensure_subscribed_callback(callback, bot):
        return

    scope = callback.data.split("_", 1)[1]  # paid/free/my
    await state.update_data(scope=scope)
    await callback.message.answer(
        "🔎 <b>Test kodini kiriting</b>\nMasalan: <code>DTM24A</code> yoki <code>MS-001</code>",
        parse_mode=ParseMode.HTML,
        reply_markup=nav_kb(include_back=True)
    )
    await state.set_state(SearchByCodeState.waiting_code)


@router.message(SearchByCodeState.waiting_code)
async def handle_code_search(message: Message, state: FSMContext):
    if (message.text or "") in TOP_MENU_BTNS:
        return

    user_db_id = await get_user_db_id(message.from_user.id)
    if not user_db_id:
        await message.answer("Avval /start orqali ro‘yxatdan o‘ting.")
        await state.clear()
        return

    st = await state.get_data()
    scope = st.get("scope", "paid")

    code = normalize_code(message.text)
    if not is_valid_code(code):
        await message.answer("⚠️ Kod formati noto‘g‘ri. Qayta kiriting:")
        return

    t = await test_by_code(code)
    if not t:
        await message.answer("❌ Bunday kodli test topilmadi. Qayta urinib ko‘ring.")
        return

    (test_id, code, title, is_free, price, duration, qcount, file_id, answers,
     start_mode, start_ts_val, start_at) = t

    scheduled_blocked, left = is_scheduled_not_started(start_mode, start_ts_val)

    # payment status (for paid)
    status = None
    payment_id = None
    started_at_paid = None
    reject_reason = None

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT p.id, p.status, p.started_at, p.reject_reason
            FROM payments p
            WHERE p.user_id=? AND p.test_id=?
            ORDER BY p.id DESC LIMIT 1
        """, (user_db_id, test_id)) as cur:
            row = await cur.fetchone()
            if row:
                payment_id, status, started_at_paid, reject_reason = row

        async with db.execute("SELECT 1 FROM results WHERE user_id=? AND test_id=? LIMIT 1", (user_db_id, test_id)) as cur:
            has_result = bool(await cur.fetchone())

    kind = "🆓 Tekin" if is_free == 1 else "🧾 Pullik"
    text = (
        f"✅ <b>Test topildi</b>\n\n"
        f"🔑 Kod: <code>{code}</code>\n"
        f"📚 Nomi: <b>{title}</b>\n"
        f"🏷 Turi: <b>{kind}</b>\n"
        f"🧮 Savollar: <b>{qcount}</b> ta\n"
        f"⏳ Vaqt: <b>{duration} daqiqa</b>\n"
    )
    if is_free == 0:
        text += f"💰 Narx: <b>{price} so'm</b>\n"

    if (start_mode or "normal") == "scheduled":
        text += f"📅 Boshlanish: <b>{start_at or 'Noma’lum'}</b>\n"
        if scheduled_blocked:
            text += f"⏳ Qolgan: <b>{fmt_seconds(left)}</b>\n"

    if is_free == 0:
        if status == "approved":
            text += "✅ Holat: Sizda bor\n"
        elif status == "pending":
            text += "🟡 Holat: To‘lov kutilmoqda\n"
        elif status == "rejected":
            text += "❌ Holat: Rad etilgan\n"
            if reject_reason:
                text += f"📝 Sabab: <i>{reject_reason}</i>\n"
        else:
            text += "ℹ️ Holat: Sotib olinmagan\n"
    else:
        if has_result:
            text += "⚠️ Siz bu tekin testni allaqachon topshirgansiz (1-urinish).\n"

    kb = InlineKeyboardMarkup(inline_keyboard=[])

    if scope == "free":
        if is_free == 1:
            if has_result:
                kb.inline_keyboard.append(
                    [InlineKeyboardButton(text="📄 Natijani ko‘rish", callback_data=f"viewres_{test_id}")])
                kb.inline_keyboard.append(
                    [InlineKeyboardButton(text="🏆 Reytingni ko‘rish", callback_data="rt_overall")])
            else:
                kb.inline_keyboard.append([InlineKeyboardButton(text="🚀 Boshlash", callback_data=f"fstart_{test_id}")])
        else:
            text += "\n⚠️ Bu test pullik. '🧾 Pullik testlar' bo‘limidan sotib oling."
    elif scope == "paid":
        if is_free == 0:
            if status == "approved":
                kb.inline_keyboard.append([InlineKeyboardButton(text="📦 Mening testlarim", callback_data="open_mytests")])
            elif status == "pending":
                kb.inline_keyboard.append([InlineKeyboardButton(text="🟡 Kutilmoqda", callback_data=f"payinfo_{test_id}")])
            elif status == "rejected":
                kb.inline_keyboard.append([InlineKeyboardButton(text="♻️ Qayta to‘lov", callback_data=f"rebuy_{test_id}")])
            else:
                kb.inline_keyboard.append([InlineKeyboardButton(text="Sotib olish", callback_data=f"buy_{test_id}")])
        else:
            text += "\n⚠️ Bu test tekin. '🆓 Tekin testlar' bo‘limidan boshlang."
    else:  # my
        if is_free == 0:
            if status == "approved":
                if has_result:
                    kb.inline_keyboard.append([InlineKeyboardButton(text="📄 Natijani ko‘rish", callback_data=f"viewres_{test_id}")])
                else:
                    kb.inline_keyboard.append([InlineKeyboardButton(text="🚀 Testni boshlash", callback_data=f"begin_{payment_id}")])
                if started_at_paid:
                    kb.inline_keyboard.append([InlineKeyboardButton(text="📄 PDF-ni ko‘rish", callback_data=f"pdf_{test_id}")])
            elif status == "pending":
                kb.inline_keyboard.append([InlineKeyboardButton(text="🟡 Pending", callback_data=f"payinfo_{test_id}")])
            elif status == "rejected":
                kb.inline_keyboard.append([InlineKeyboardButton(text="♻️ Qayta to‘lov yuborish", callback_data=f"rebuy_{test_id}")])
            else:
                kb.inline_keyboard.append([InlineKeyboardButton(text="Sotib olish", callback_data=f"buy_{test_id}")])
        else:
            kb.inline_keyboard.append([InlineKeyboardButton(text="🆓 Tekin testlar bo‘limiga o‘ting", callback_data="noop")])

    await message.answer(text, parse_mode=ParseMode.HTML, reply_markup=kb if kb.inline_keyboard else None)
    await state.clear()


@router.callback_query(F.data == "noop")
async def noop(callback: CallbackQuery):
    await callback.answer("OK", show_alert=False)


# ===================== FREE TESTS =====================
@router.message(F.text == BTN_FREE)
async def show_free_tests(message: Message):
    user_db_id = await get_user_db_id(message.from_user.id)
    if not user_db_id:
        return await message.answer("Avval /start orqali ro‘yxatdan o‘ting.")

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT id, code, title, duration, questions_count, start_mode, start_ts, start_at
            FROM tests
            WHERE is_free=1
            ORDER BY id DESC
            LIMIT 30
        """) as cur:
            tests = await cur.fetchall()

    if not tests:
        return await message.answer("Hozircha tekin testlar yo‘q.", reply_markup=search_kb("free"))

    text = "<b>🆓 Tekin testlar (oxirgi 30 ta):</b>\n\n"
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    kb.inline_keyboard.append([InlineKeyboardButton(text="🔎 Kod orqali topish", callback_data="askcode_free")])

    for tid, code, title, duration, qcount, start_mode, start_ts_val, start_at in tests:
        sched = (start_mode or "normal") == "scheduled"
        blocked, left = is_scheduled_not_started(start_mode, start_ts_val)

        line = f"🔑 <code>{code}</code> — <b>{title}</b> | 🧮 {qcount} ta | ⏳ {duration} daqiqa"
        if sched:
            line += f"\n   📅 Boshlanish: <b>{start_at}</b>"
            if blocked:
                line += f" | ⏳ {fmt_seconds(left)}"
        text += line + "\n\n"

        kb.inline_keyboard.append([InlineKeyboardButton(text=f"🚀 Boshlash: {code}", callback_data=f"fstart_{tid}")])

    await message.answer(text, parse_mode=ParseMode.HTML, reply_markup=kb)


@router.callback_query(F.data.startswith("fstart_"))
async def free_test_start(callback: CallbackQuery, state: FSMContext, bot: Bot):
    if not await ensure_subscribed_callback(callback, bot):
        return

    test_id = int(callback.data.split("_")[1])

    user_db_id = await get_user_db_id(callback.from_user.id)
    if not user_db_id:
        return await callback.message.answer("Avval /start orqali ro‘yxatdan o‘ting.")

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT 1 FROM results WHERE user_id=? AND test_id=? LIMIT 1", (user_db_id, test_id)) as cur:
            if await cur.fetchone():
                return await callback.message.answer("⚠️ Siz bu testni allaqachon topshirgansiz. Qayta topshirib bo‘lmaydi.")

        async with db.execute("""
            SELECT code, title, file_id, duration, questions_count, answers, start_mode, start_ts, start_at
            FROM tests
            WHERE id=? AND is_free=1
        """, (test_id,)) as cur:
            row = await cur.fetchone()

    if not row:
        return await callback.message.answer("Test topilmadi (yoki tekin emas).")

    code, title, file_id, duration_min, qcount, answers, start_mode, start_ts_val, start_at = row

    blocked, left = is_scheduled_not_started(start_mode, start_ts_val)
    if blocked:
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔁 Qayta tekshirish", callback_data=f"chkfree_{test_id}")],
        ])
        return await callback.message.answer(
            f"⏳ <b>Test hali boshlanmagan</b>\n\n"
            f"🔑 Kod: <code>{code}</code>\n"
            f"📅 Boshlanish: <b>{start_at}</b>\n"
            f"⏳ Qolgan: <b>{fmt_seconds(left)}</b>\n\n"
            f"Iltimos, kuting.",
            parse_mode=ParseMode.HTML,
            reply_markup=kb
        )

    answer_key = normalize_answers(answers)
    if len(answer_key) != int(qcount):
        return await callback.message.answer("⚠️ Admin testni xato qo‘shgan: javob kaliti uzunligi savol soniga teng emas.")

    cancel_timer(callback.from_user.id, test_id)

    start_ts_ = int(start_ts_val) if (start_mode or "normal") == "scheduled" and int(start_ts_val or 0) > 0 else now_ts()
    duration_sec = int(duration_min) * 60
    end_ts_ = start_ts_ + duration_sec

    if now_ts() > end_ts_ + 60:
        return await callback.message.answer("❌ Bu test vaqti tugab bo‘lgan.")

    await upsert_session(user_db_id, test_id, "free", start_ts_, duration_sec)

    await bot.send_document(
        chat_id=callback.from_user.id,
        document=file_id,
        caption=(
            f"📄 <b>{title}</b>\n"
            f"🔑 Kod: <code>{code}</code>\n"
            f"🧮 Savollar: <b>{qcount}</b>\n\n"
            f"🆓 Tekin test. Vaqt ketdi ✅"
        ),
        parse_mode=ParseMode.HTML
    )

    timer_msg = await bot.send_message(
        chat_id=callback.from_user.id,
        text="⏳ <b>Qolgan vaqt:</b> ...",
        parse_mode=ParseMode.HTML
    )

    ACTIVE_TIMERS[(callback.from_user.id, test_id)] = asyncio.create_task(
        start_countdown(bot, callback.from_user.id, test_id, end_ts_, timer_msg.message_id)
    )

    await state.update_data(test_id=test_id, mode="free")
    await bot.send_message(
        chat_id=callback.from_user.id,
        text=(
            f"⏱ <b>Test boshlandi!</b>\n"
            f"🔑 Kod: <code>{code}</code>\n"
            f"🧮 Savollar: <b>{qcount}</b>\n\n"
            f"Javob format: <code>abcd...</code> (bo‘sh joysiz)\n"
            f"⚠️ Faqat 1-urinish qabul qilinadi.\n"
            f"⚠️ Javoblar soni aniq <b>{qcount}</b> ta bo‘lishi shart."
        ),
        parse_mode=ParseMode.HTML,
        reply_markup=nav_kb(include_back=True)
    )
    await state.set_state(TestProcessState.solving)


@router.callback_query(F.data.startswith("chkfree_"))
async def check_free_start(callback: CallbackQuery):
    test_id = int(callback.data.split("_")[1])
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT code, start_mode, start_ts, start_at FROM tests WHERE id=?", (test_id,)) as cur:
            row = await cur.fetchone()
    if not row:
        return await callback.message.answer("Test topilmadi.")
    code, start_mode, start_ts_val, start_at = row
    blocked, left = is_scheduled_not_started(start_mode, start_ts_val)
    if blocked:
        return await callback.message.answer(
            f"⏳ Hali erta.\n🔑 <code>{code}</code>\n📅 {start_at}\n⏳ Qolgan: {fmt_seconds(left)}",
            parse_mode=ParseMode.HTML
        )
    await callback.message.answer("✅ Boshlash vaqti keldi! Endi 'Boshlash' tugmasini bosing.")


# ===================== PAID TESTS =====================
@router.message(F.text == BTN_PAID)
async def show_paid_tests(message: Message):
    user_db_id = await get_user_db_id(message.from_user.id)
    if not user_db_id:
        return await message.answer("Avval /start orqali ro‘yxatdan o‘ting.")

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT id, code, title, price, duration, questions_count, start_mode, start_ts, start_at
            FROM tests
            WHERE is_free=0
            ORDER BY id DESC
            LIMIT 30
        """) as cur:
            tests = await cur.fetchall()

        async with db.execute("""
            SELECT p.test_id, p.status
            FROM payments p
            WHERE p.user_id=?
            ORDER BY p.id DESC
        """, (user_db_id,)) as cur:
            pay_rows = await cur.fetchall()

    if not tests:
        return await message.answer("Hozircha pullik testlar yo‘q.", reply_markup=search_kb("paid"))

    latest_status = {}
    for tid, st in pay_rows:
        if tid not in latest_status:
            latest_status[tid] = st

    text = "<b>🧾 Pullik testlar (oxirgi 30 ta):</b>\n\n"
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    kb.inline_keyboard.append([InlineKeyboardButton(text="🔎 Kod orqali topish", callback_data="askcode_paid")])

    for tid, code, title, price, duration, qcount, start_mode, start_ts_val, start_at in tests:
        st = latest_status.get(tid)
        st_txt = ""
        if st == "approved":
            st_txt = "✅ Sizda bor"
        elif st == "pending":
            st_txt = "🟡 Kutilmoqda"
        elif st == "rejected":
            st_txt = "❌ Rad"

        sched = (start_mode or "normal") == "scheduled"
        blocked, left = is_scheduled_not_started(start_mode, start_ts_val)

        text += f"🔑 <code>{code}</code> — <b>{title}</b>\n"
        text += f"🧮 {qcount} ta | 💰 {price} | ⏳ {duration} daq"
        if st_txt:
            text += f" | {st_txt}"
        text += "\n"
        if sched:
            text += f"📅 Boshlanish: <b>{start_at}</b>"
            if blocked:
                text += f" | ⏳ {fmt_seconds(left)}"
            text += "\n"
        text += "\n"

        if st == "approved":
            kb.inline_keyboard.append([InlineKeyboardButton(text=f"📦 Mening testlarim: {code}", callback_data="open_mytests")])
        elif st == "pending":
            kb.inline_keyboard.append([InlineKeyboardButton(text=f"🟡 Pending: {code}", callback_data=f"payinfo_{tid}")])
        elif st == "rejected":
            kb.inline_keyboard.append([InlineKeyboardButton(text=f"♻️ Qayta to'lov: {code}", callback_data=f"rebuy_{tid}")])
        else:
            kb.inline_keyboard.append([InlineKeyboardButton(text=f"Sotib olish: {code}", callback_data=f"buy_{tid}")])

    await message.answer(text, parse_mode=ParseMode.HTML, reply_markup=kb)


@router.callback_query(F.data == "open_mytests")
async def open_mytests_cb(callback: CallbackQuery):
    await callback.message.answer("📦 Mening testlarim bo‘limiga o‘ting (menyudan).")


@router.callback_query(F.data.startswith("payinfo_"))
async def payinfo(callback: CallbackQuery):
    await callback.message.answer("🟡 Bu test bo‘yicha to‘lov admin tasdiqlashini kutyapti.")


@router.callback_query(F.data.startswith("buy_"))
async def start_payment(callback: CallbackQuery, state: FSMContext, bot: Bot):
    if not await ensure_subscribed_callback(callback, bot):
        return

    user_db_id = await get_user_db_id(callback.from_user.id)
    if not user_db_id:
        return await callback.message.answer("Avval /start orqali ro‘yxatdan o‘ting.")

    test_id = int(callback.data.split("_")[1])
    # ✅ TEST EXPIRED CHECK (to‘lovdan oldin!)
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT code, title, start_mode, start_ts, start_at, duration
            FROM tests
            WHERE id=? AND is_free=0
        """, (test_id,)) as cur:
            tr = await cur.fetchone()

    if not tr:
        return await callback.message.answer("Test topilmadi.")

    tcode, ttitle, tmode, tstart_ts, tstart_at, tdur = tr

    expired, after = is_scheduled_expired(tmode, tstart_ts, tdur)
    if expired:
        return await callback.message.answer(
            f"❌ <b>Bu pullik testning vaqti tugab bo‘lgan.</b>\n\n"
            f"🔑 Kod: <code>{tcode}</code>\n"
            f"📚 Test: <b>{ttitle}</b>\n"
            f"📅 Boshlanish: <b>{tstart_at or 'Noma’lum'}</b>\n\n"
            f"✅ To‘lov qabul qilinmaydi.",
            parse_mode=ParseMode.HTML
        )

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT 1 FROM payments
            WHERE user_id=? AND test_id=? AND status='pending'
            LIMIT 1
        """, (user_db_id, test_id)) as cur:
            if await cur.fetchone():
                return await callback.message.answer("🟡 Sizda shu test bo‘yicha pending to‘lov bor. Admin javobini kuting.")

    await state.update_data(test_id=test_id)
    await callback.message.answer(
        "💳 <b>To'lov ma'lumotlari:</b>\n\n"
        f"💳 Karta: <code>{PAYMENT_CARD}</code>\n"
        f"👤 Ega: <b>{PAYMENT_OWNER}</b>\n"
        f"📞 Telefon: <code>{PAYMENT_PHONE}</code>\n"
        f"📝 Izoh: <i>{PAYMENT_NOTE}</i>\n\n"
        "✅ To'lov qilib bo‘lgach, chek (skrinshot) yuboring. Va tezroq tasdiqlanishi uchun berilgan raqamga telefon qiling",
        parse_mode=ParseMode.HTML,
        reply_markup=nav_kb(include_back=True)
    )
    await state.set_state(PaymentState.waiting_for_screenshot)


@router.callback_query(F.data.startswith("rebuy_"))
async def rebuy(callback: CallbackQuery, state: FSMContext, bot: Bot):
    if not await ensure_subscribed_callback(callback, bot):
        return

    user_db_id = await get_user_db_id(callback.from_user.id)
    if not user_db_id:
        return await callback.message.answer("Avval /start orqali ro‘yxatdan o‘ting.")

    test_id = int(callback.data.split("_")[1])
    # ✅ TEST EXPIRED CHECK (to‘lovdan oldin!)
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT code, title, start_mode, start_ts, start_at, duration
            FROM tests
            WHERE id=? AND is_free=0
        """, (test_id,)) as cur:
            tr = await cur.fetchone()

    if not tr:
        return await callback.message.answer("Test topilmadi.")

    tcode, ttitle, tmode, tstart_ts, tstart_at, tdur = tr

    expired, after = is_scheduled_expired(tmode, tstart_ts, tdur)
    if expired:
        return await callback.message.answer(
            f"❌ <b>Bu pullik testning vaqti tugab bo‘lgan.</b>\n\n"
            f"🔑 Kod: <code>{tcode}</code>\n"
            f"📚 Test: <b>{ttitle}</b>\n"
            f"📅 Boshlanish: <b>{tstart_at or 'Noma’lum'}</b>\n\n"
            f"✅ To‘lov qabul qilinmaydi.",
            parse_mode=ParseMode.HTML
        )

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT 1 FROM payments
            WHERE user_id=? AND test_id=? AND status='pending'
            LIMIT 1
        """, (user_db_id, test_id)) as cur:
            if await cur.fetchone():
                return await callback.message.answer("🟡 Sizda shu test bo‘yicha pending to‘lov bor. Admin javobini kuting.")

    await state.update_data(test_id=test_id)
    await callback.message.answer(
        "💳 <b>To'lov ma'lumotlari:</b>\n\n"
        f"💳 Karta: <code>{PAYMENT_CARD}</code>\n"
        f"👤 Ega: <b>{PAYMENT_OWNER}</b>\n"
        f"📞 Telefon: <code>{PAYMENT_PHONE}</code>\n"
        f"📝 Izoh: <i>{PAYMENT_NOTE}</i>\n\n"
        "✅ To'lov qilib bo‘lgach, chek (skrinshot) yuboring. Va tezroq tasdiqlanishi uchun berilgan raqamga telefon qiling",
        parse_mode=ParseMode.HTML,
        reply_markup=nav_kb(include_back=True)
    )
    await state.set_state(PaymentState.waiting_for_screenshot)


@router.message(PaymentState.waiting_for_screenshot, F.photo)
async def process_screenshot(message: Message, state: FSMContext, bot: Bot):
    data = await state.get_data()
    test_id = int(data.get("test_id") or 0)
    photo_id = message.photo[-1].file_id

    user_db_id = await get_user_db_id(message.from_user.id)
    if not user_db_id:
        await message.answer("Avval /start orqali ro‘yxatdan o‘ting.")
        await state.clear()
        return
    # ✅ EXPIRED CHECK (skrin qabul qilishdan oldin)
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT code, title, start_mode, start_ts, start_at, duration
            FROM tests
            WHERE id=? AND is_free=0
        """, (test_id,)) as cur:
            tr = await cur.fetchone()

    if not tr:
        await message.answer("Test topilmadi.")
        await state.clear()
        return

    tcode, ttitle, tmode, tstart_ts, tstart_at, tdur = tr
    expired, after = is_scheduled_expired(tmode, tstart_ts, tdur)
    if expired:
        await message.answer(
            f"❌ <b>Bu testning vaqti tugab bo‘lgan.</b>\n\n"
            f"🔑 Kod: <code>{tcode}</code>\n"
            f"📚 Test: <b>{ttitle}</b>\n"
            f"📅 Boshlanish: <b>{tstart_at or 'Noma’lum'}</b>\n\n"
            f"✅ Chek qabul qilinmaydi.",
            parse_mode=ParseMode.HTML
        )
        await state.clear()
        return

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT 1 FROM payments
            WHERE user_id=? AND test_id=? AND status='pending'
            LIMIT 1
        """, (user_db_id, test_id)) as cur:
            if await cur.fetchone():
                await message.answer("🟡 Sizda shu test bo‘yicha pending to‘lov bor.")
                await state.clear()
                return


        cur = await db.execute("""
            INSERT INTO payments (user_id, test_id, screenshot_id, status, created_at)
            VALUES (?, ?, ?, 'pending', ?)
        """, (user_db_id, test_id, photo_id, now_str_local()))
        payment_id = cur.lastrowid
        await db.commit()

        async with db.execute("SELECT full_name FROM users WHERE id=?", (user_db_id,)) as ucur:
            uname = (await ucur.fetchone())[0]

        async with db.execute("SELECT code, title, price FROM tests WHERE id=?", (test_id,)) as tcur:
            t = await tcur.fetchone()
            if not t:
                await message.answer("Test topilmadi.")
                await state.clear()
                return
            code, title, price = t

    admin_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Tasdiqlash", callback_data=f"appr_{payment_id}")],
        [InlineKeyboardButton(text="❌ Rad etish (Sabab)", callback_data=f"rejask_{payment_id}")]
    ])

    await bot.send_photo(
        chat_id=ADMIN_ID,
        photo=photo_id,
        caption=(
            f"🟡 <b>Yangi to'lov (PENDING)</b>\n\n"
            f"👤 User: {uname}\n"
            f"📚 Test: {title}\n"
            f"🔑 Kod: <code>{code}</code>\n"
            f"💰 Summa: {price} so'm\n"
            f"🆔 Payment ID: <code>{payment_id}</code>"
        ),
        parse_mode=ParseMode.HTML,
        reply_markup=admin_kb
    )

    admin_flag = await is_admin_user(message.from_user.id)
    await message.answer("✅ Chek qabul qilindi. Admin tasdiqlashini kuting.", reply_markup=main_menu_kb(admin_flag))
    await state.clear()


# ===================== MY TESTS =====================
@router.message(F.text == BTN_MY)
async def my_tests(message: Message):
    user_db_id = await get_user_db_id(message.from_user.id)
    if not user_db_id:
        return await message.answer("Avval /start orqali ro‘yxatdan o‘ting.")

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT
                p.id AS payment_id,
                p.status,
                p.test_id,
                p.reject_reason,
                p.started_at,
                t.code,
                t.title,
                t.duration,
                t.price,
                t.questions_count,
                t.start_mode,
                t.start_ts,
                t.start_at,
                (
                    SELECT r.percent
                    FROM results r
                    WHERE r.user_id=p.user_id AND r.test_id=p.test_id
                    ORDER BY r.id ASC LIMIT 1
                ) AS first_percent
            FROM payments p
            JOIN tests t ON t.id=p.test_id
            WHERE p.user_id=?
            ORDER BY p.id DESC
            LIMIT 200
        """, (user_db_id,)) as cur:
            rows = await cur.fetchall()

    actions = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🔎 Kod orqali topish", callback_data="askcode_my")]
    ])

    if not rows:
        admin_flag = await is_admin_user(message.from_user.id)
        await message.answer("Siz hali pullik test sotib olmagansiz.", reply_markup=main_menu_kb(admin_flag))
        await message.answer("Kod orqali qidirish:", reply_markup=actions)
        return

    seen = set()
    items = []
    for r in rows:
        (payment_id, status, test_id, reject_reason, started_at_paid, code, title,
         duration, price, qcount, start_mode, start_ts_val, start_at, first_percent) = r
        if test_id in seen:
            continue
        seen.add(test_id)
        items.append(r)

    text = "<b>📦 Mening pullik testlarim</b>\n\n"
    for r in items:
        (payment_id, status, test_id, reject_reason, started_at_paid, code, title,
         duration, price, qcount, start_mode, start_ts_val, start_at, first_percent) = r

        sched = (start_mode or "normal") == "scheduled"
        blocked, left = is_scheduled_not_started(start_mode, start_ts_val)

        text += f"🔑 <code>{code}</code> — <b>{title}</b>\n"
        text += f"🧮 {qcount} ta | 💰 {price} | ⏳ {duration} daq\n"
        if sched:
            text += f"📅 Boshlanish: <b>{start_at}</b>"
            if blocked:
                text += f" | ⏳ {fmt_seconds(left)}"
            text += "\n"

        if status == "approved":
            if first_percent is not None:
                text += f"✅ Natija: <b>{float(first_percent):.1f}%</b> (1-urinish)\n\n"
                actions.inline_keyboard.append([InlineKeyboardButton(text=f"📄 Natija: {code}", callback_data=f"viewres_{test_id}")])
                if started_at_paid:
                    actions.inline_keyboard.append([InlineKeyboardButton(text=f"📄 PDF: {code}", callback_data=f"pdf_{test_id}")])
            else:
                text += "✅ Tasdiqlangan (hali yechilmagan)\n\n"
                actions.inline_keyboard.append([InlineKeyboardButton(text=f"🚀 Boshlash: {code}", callback_data=f"begin_{payment_id}")])

        elif status == "pending":
            text += "🟡 Pending\n\n"
        elif status == "rejected":
            text += "❌ Rad etilgan"
            if reject_reason:
                text += f" — {reject_reason}"
            text += "\n\n"
            actions.inline_keyboard.append([InlineKeyboardButton(text=f"♻️ Qayta to‘lov: {code}", callback_data=f"rebuy_{test_id}")])
        else:
            text += "ℹ️ Noma’lum holat\n\n"

    await message.answer(text, parse_mode=ParseMode.HTML, reply_markup=actions)


@router.callback_query(F.data.startswith("pdf_"))
async def send_paid_pdf(callback: CallbackQuery, bot: Bot):
    test_id = int(callback.data.split("_")[1])
    user_db_id = await get_user_db_id(callback.from_user.id)
    if not user_db_id:
        return await callback.message.answer("Avval /start orqali ro‘yxatdan o‘ting.")

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT t.code, t.title, t.file_id, p.started_at
            FROM payments p
            JOIN tests t ON t.id=p.test_id
            WHERE p.user_id=? AND p.test_id=? AND p.status='approved'
            ORDER BY p.id DESC LIMIT 1
        """, (user_db_id, test_id)) as cur:
            row = await cur.fetchone()

    if not row:
        return await callback.message.answer("Bu test sizda topilmadi yoki tasdiqlanmagan.")

    code, title, file_id, started_at = row
    if not started_at:
        return await callback.message.answer("⚠️ PDFni ko‘rish uchun avval testni 'Boshlash' tugmasi orqali boshlang.")

    await bot.send_document(
        chat_id=callback.from_user.id,
        document=file_id,
        caption=f"📄 <b>{title}</b>\n🔑 Kod: <code>{code}</code>",
        parse_mode=ParseMode.HTML
    )


@router.callback_query(F.data.startswith("viewres_"))
async def view_result(callback: CallbackQuery):
    test_id = int(callback.data.split("_")[1])
    user_db_id = await get_user_db_id(callback.from_user.id)
    if not user_db_id:
        return await callback.message.answer("Avval /start orqali ro‘yxatdan o‘ting.")

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT t.code, t.title, t.subject,t.is_free, t.price, t.exam_type,
                   r.score, r.total_questions, r.percent,
                   r.earned_points, r.max_points, r.details,
                   r.rasch_ball, r.rasch_percent, r.grade, r.date
            FROM results r
            JOIN tests t ON t.id=r.test_id
            WHERE r.user_id=? AND r.test_id=?
            ORDER BY r.id ASC LIMIT 1
        """, (user_db_id, test_id)) as cur:
            row = await cur.fetchone()

    if not row:
        return await callback.message.answer("Bu test bo‘yicha natija topilmadi.")

    (code, title, subject,is_free, price, exam_type,
     score, total, percent,
     earned_points, max_points, details_json,
     rasch_ball, rasch_percent, grade, date_) = row

    exam_type = (exam_type or "simple")

    # ✅ finalize: rasch / maxsus
    if exam_type == "rasch":
        await finalize_rush_for_test_if_ready(test_id)

        async with aiosqlite.connect(DB_NAME) as db:
            async with db.execute("""
                SELECT rasch_ball, rasch_percent, grade
                FROM results
                WHERE user_id=? AND test_id=?
                ORDER BY id ASC LIMIT 1
            """, (user_db_id, test_id)) as cur:
                rr = await cur.fetchone()
                if rr:
                    rasch_ball, rasch_percent, grade = rr

    elif exam_type == "maxsus":
        await finalize_maxsus_for_test_if_ready(test_id)

        async with aiosqlite.connect(DB_NAME) as db:
            async with db.execute("""
                SELECT rasch_ball, rasch_percent, grade
                FROM results
                WHERE user_id=? AND test_id=?
                ORDER BY id ASC LIMIT 1
            """, (user_db_id, test_id)) as cur:
                rr = await cur.fetchone()
                if rr:
                    rasch_ball, rasch_percent, grade = rr

        # finalize dan keyin rasch ustunlari yangilangan bo'lishi mumkin -> qayta o'qib olamiz
        async with aiosqlite.connect(DB_NAME) as db:
            async with db.execute("""
                SELECT rasch_ball, rasch_percent, grade
                FROM results
                WHERE user_id=? AND test_id=?
                ORDER BY id ASC LIMIT 1
            """, (user_db_id, test_id)) as cur:
                rr = await cur.fetchone()
                if rr:
                    rasch_ball, rasch_percent, grade = rr

    pay_line = ""
    if int(is_free or 0) == 0:
        pay_line = f"💳 To‘lov: <b>{int(price or 0)} so‘m</b>\n"

    # ✅ ODDIY blok
    base_block = (
        f"📄 <b>Natija (1-urinish)</b>\n\n"
        f"🔑 Kod: <code>{code}</code>\n"
        f"📚 Fan: <b>{subject or 'Noma’lum'}</b>\n"
        f"📚 Test: <b>{title}</b>\n"
        f"{pay_line}"
        f"✅ To‘g‘ri: {score}/{total}\n"
        f"❌ Xato: {int(total or 0) - int(score or 0)}\n"
        f"📊 Oddiy foiz: <b>{float(percent):.1f}%</b>\n"
    )

    # ✅ DTM blok
    dtm_block = ""
    if exam_type == "dtm":
        dtm_block = (
            f"\n🏁 Ball: <b>{float(earned_points or 0):.1f}/{float(max_points or 0):.1f}</b>\n"
            f"📌 Ball foiz: <b>{float(percent):.1f}%</b>\n"
        )

    # ✅ RASCH blok (faqat rasch)
    rasch_block = ""
    if exam_type in ("rasch", "maxsus"):
        if rasch_ball is None:
            rasch_block = (
                "\n⚡ Rasch natija: <b>Hali hisoblanmadi (test tugagach chiqadi)</b>\n"
                "🎯 Daraja: <b>Test tugagach chiqadi</b>\n"
            )
        else:
            if grade is None:
                rasch_block = (
                    f"\n⚡ Rasch ball: <b>{float(rasch_ball):.1f} / 75</b>\n"
                    f"⚡ Rasch foiz: <b>{float(rasch_percent or 0):.1f}%</b>\n"
                    f"🎯 Daraja: <b>Sertifikat yo‘q</b>\n"
                )
            else:
                rasch_block = (
                    f"\n⚡ Rasch ball: <b>{float(rasch_ball):.1f} / 75</b>\n"
                    f"⚡ Rasch foiz: <b>{float(rasch_percent or 0):.1f}%</b>\n"
                    f"🎯 Daraja: <b>{grade}</b>\n"
                )

    await callback.message.answer(
        base_block + dtm_block + rasch_block + f"\n🗓 Sana: {date_}",
        parse_mode=ParseMode.HTML
    )



# ===================== PAID TEST BEGIN =====================
@router.callback_query(F.data.startswith("begin_"))
async def begin_test(callback: CallbackQuery, state: FSMContext, bot: Bot):
    if not await ensure_subscribed_callback(callback, bot):
        return

    payment_id = int(callback.data.split("_")[1])
    user_db_id = await get_user_db_id(callback.from_user.id)
    if not user_db_id:
        return await callback.message.answer("Avval /start orqali ro‘yxatdan o‘ting.")

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT p.status, p.started_at, p.started_ts, p.test_id,
                   t.code, t.title, t.file_id, t.duration, t.questions_count, t.answers,
                   t.start_mode, t.start_ts, t.start_at
            FROM payments p
            JOIN tests t ON t.id=p.test_id
            WHERE p.id=? AND p.user_id=?
        """, (payment_id, user_db_id)) as cur:
            row = await cur.fetchone()

    if not row:
        return await callback.message.answer("Xatolik: payment topilmadi.")

    (status, started_at_paid, started_ts_paid, test_id,
     code, title, file_id, duration_min, qcount, answers,
     start_mode, start_ts_val, start_at) = row

    if status != "approved":
        return await callback.message.answer("Bu test hali tasdiqlanmagan.")

    blocked, left = is_scheduled_not_started(start_mode, start_ts_val)
    if blocked:
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔁 Qayta tekshirish", callback_data=f"chkpaid_{payment_id}")]
        ])
        return await callback.message.answer(
            f"⏳ <b>Test hali boshlanmagan</b>\n\n"
            f"🔑 Kod: <code>{code}</code>\n"
            f"📅 Boshlanish: <b>{start_at}</b>\n"
            f"⏳ Qolgan: <b>{fmt_seconds(left)}</b>\n\n"
            f"Iltimos, kuting.",
            parse_mode=ParseMode.HTML,
            reply_markup=kb
        )

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT 1 FROM results WHERE user_id=? AND test_id=? LIMIT 1", (user_db_id, test_id)) as cur:
            if await cur.fetchone():
                return await callback.message.answer("⚠️ Siz bu testni allaqachon topshirgansiz. Qayta topshirib bo‘lmaydi.")

    answer_key = normalize_answers(answers)
    if len(answer_key) != int(qcount):
        return await callback.message.answer("⚠️ Admin testni xato qo‘shgan: javob kaliti uzunligi savol soniga teng emas.")

    cancel_timer(callback.from_user.id, test_id)

    exam_start_ts = int(start_ts_val) if (start_mode or "normal") == "scheduled" and int(start_ts_val or 0) > 0 else now_ts()
    duration_sec = int(duration_min) * 60
    end_ts_ = exam_start_ts + duration_sec

    if now_ts() > end_ts_ + 60:
        return await callback.message.answer("❌ Bu test vaqti tugab bo‘lgan.")

    if not started_ts_paid:
        async with aiosqlite.connect(DB_NAME) as db:
            await db.execute("UPDATE payments SET started_at=?, started_ts=? WHERE id=?", (now_str_local(), now_ts(), payment_id))
            await db.commit()

    await upsert_session(user_db_id, test_id, "paid", exam_start_ts, duration_sec)

    await bot.send_document(
        chat_id=callback.from_user.id,
        document=file_id,
        caption=(
            f"📄 <b>{title}</b>\n"
            f"🔑 Kod: <code>{code}</code>\n"
            f"🧮 Savollar: <b>{qcount}</b>\n\n"
            f"✅ Test boshlandi. Vaqt ketdi!"
        ),
        parse_mode=ParseMode.HTML
    )

    timer_msg = await bot.send_message(
        chat_id=callback.from_user.id,
        text="⏳ <b>Qolgan vaqt:</b> ...",
        parse_mode=ParseMode.HTML
    )

    ACTIVE_TIMERS[(callback.from_user.id, test_id)] = asyncio.create_task(
        start_countdown(bot, callback.from_user.id, test_id, end_ts_, timer_msg.message_id)
    )

    await state.update_data(test_id=test_id, mode="paid")
    await bot.send_message(
        chat_id=callback.from_user.id,
        text=(
            f"⏱ <b>Javoblarni yuboring</b>\n\n"
            f"🔑 Kod: <code>{code}</code>\n"
            f"🧮 Savollar: <b>{qcount}</b>\n\n"
            f"Javob format: <code>abcd...</code>\n"
            f"⚠️ Faqat 1-urinish.\n"
            f"⚠️ Javoblar soni aniq <b>{qcount}</b> ta bo‘lishi shart."
        ),
        parse_mode=ParseMode.HTML,
        reply_markup=nav_kb(include_back=True)
    )
    await state.set_state(TestProcessState.solving)


@router.callback_query(F.data.startswith("chkpaid_"))
async def check_paid_start(callback: CallbackQuery):
    payment_id = int(callback.data.split("_")[1])
    user_db_id = await get_user_db_id(callback.from_user.id)
    if not user_db_id:
        return await callback.message.answer("Avval /start orqali ro‘yxatdan o‘ting.")

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT t.code, t.start_mode, t.start_ts, t.start_at
            FROM payments p
            JOIN tests t ON t.id=p.test_id
            WHERE p.id=? AND p.user_id=?
        """, (payment_id, user_db_id)) as cur:
            row = await cur.fetchone()

    if not row:
        return await callback.message.answer("Topilmadi.")

    code, start_mode, start_ts_val, start_at = row
    blocked, left = is_scheduled_not_started(start_mode, start_ts_val)
    if blocked:
        return await callback.message.answer(
            f"⏳ Hali erta.\n🔑 <code>{code}</code>\n📅 {start_at}\n⏳ Qolgan: {fmt_seconds(left)}",
            parse_mode=ParseMode.HTML
        )
    await callback.message.answer("✅ Boshlash vaqti keldi! Endi 'Boshlash' tugmasini bosing.")


# ===================== SUBMIT ANSWERS (FREE+PAID) =====================
@router.message(TestProcessState.solving)
async def submit_answers(message: Message, state: FSMContext):
    if (message.text or "") in TOP_MENU_BTNS:
        return

    user_db_id = await get_user_db_id(message.from_user.id)
    if not user_db_id:
        await message.answer("Avval /start orqali ro‘yxatdan o‘ting.")
        await state.clear()
        return

    st = await state.get_data()
    test_id = int(st.get("test_id") or 0)
    mode = st.get("mode") or "free"

    sess = await get_session(user_db_id, test_id, mode)
    if not sess:
        await message.answer("⚠️ Sessiya topilmadi. Testni qayta boshlang.")
        cancel_timer(message.from_user.id, test_id)
        await state.clear()
        return

    start_ts_, duration_sec = sess
    end_ts_ = int(start_ts_) + int(duration_sec)
    if now_ts() > end_ts_ + 60:
        await message.answer("❌ Vaqt tugadi! Javoblar qabul qilinmadi.")
        await delete_session(user_db_id, test_id, mode)
        cancel_timer(message.from_user.id, test_id)
        await state.clear()
        return

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT 1 FROM results WHERE user_id=? AND test_id=? LIMIT 1", (user_db_id, test_id)) as cur:
            if await cur.fetchone():
                await message.answer("⚠️ Siz bu testni allaqachon topshirgansiz. Qayta topshirib bo‘lmaydi.")
                await delete_session(user_db_id, test_id, mode)
                cancel_timer(message.from_user.id, test_id)
                await state.clear()
                return

        async with db.execute("""
            SELECT code, title, subject, questions_count, answers, exam_type, dtm_cfg, start_mode, start_ts, duration
            FROM tests
            WHERE id=?

        """, (test_id,)) as cur:
            t = await cur.fetchone()

    if not t:
        await message.answer("Test topilmadi.")
        await delete_session(user_db_id, test_id, mode)
        cancel_timer(message.from_user.id, test_id)
        await state.clear()
        return

    code, title, subject, qcount, answer_key_raw, exam_type, dtm_cfg, start_mode, start_ts_val, duration_min = t

    qcount = int(qcount)
    correct_key = normalize_answers(answer_key_raw)

    if len(correct_key) != qcount:
        await message.answer("⚠️ Test javob kaliti xato. Admin bilan bog‘laning.")
        await delete_session(user_db_id, test_id, mode)
        cancel_timer(message.from_user.id, test_id)
        await state.clear()
        return

    user_ans = normalize_answers(message.text)

    if not is_answer_string_valid(user_ans):
        await message.answer("⚠️ Javoblar faqat harflardan iborat bo‘lsin. Masalan: <code>abcd...</code>", parse_mode=ParseMode.HTML)
        return

    if len(user_ans) != qcount:
        await message.answer(
            f"⚠️ Javoblar soni noto‘g‘ri.\n"
            f"Kerakli: <b>{qcount}</b>, Siz yubordingiz: <b>{len(user_ans)}</b>.\n\n"
            f"Iltimos, aniq <b>{qcount} ta</b> javob yuboring.",
            parse_mode=ParseMode.HTML
        )
        return

    score = 0
    mistakes = []
    for i in range(qcount):
        if user_ans[i] == correct_key[i]:
            score += 1
        else:
            mistakes.append(f"{i+1}-savol: siz={user_ans[i]} | to‘g‘ri={correct_key[i]}")

    percent = (score / qcount) * 100

    earned_points = float(score)
    max_points = float(qcount)
    details = None

    # ✅ DTM bo'lsa ball bo'yicha hisoblaymiz
    if (exam_type or "simple") == "dtm":
        earned_points, max_points, details_obj = dtm_score_points(user_ans, correct_key, dtm_cfg or "")
        details = json.dumps(details_obj, ensure_ascii=False)
        percent_points = (earned_points / max_points) * 100 if max_points > 0 else 0.0
    else:
        percent_points = percent  # oddiyda bir xil

    async with aiosqlite.connect(DB_NAME) as db:
        await db.execute("""
            INSERT INTO results (
                user_id, test_id, score, total_questions, percent,
                user_answers,
                earned_points, max_points, details,
                rasch_ball, rasch_percent, grade,
                finished_ts, date
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, NULL, NULL, NULL, ?, ?)
        """, (
            user_db_id, test_id,
            score, qcount,
            float(percent_points),  # ✅ DTMda ball foiz, oddiyda oddiy foiz
            user_ans,
            float(earned_points), float(max_points), details,
            now_ts(),
            now_str_local()
        ))

        await db.commit()
    # ✅ CERTIFICATE: immediate only for normal + (simple/dtm) and eligible
    et = (exam_type or "simple")
    can_immediate = ((start_mode or "normal") == "normal") and (et in ("simple", "dtm"))
    if can_immediate:
        eligible = is_certificate_eligible(et, percent_points, None)
        if eligible:
            try:
                user_full_name = await get_user_full_name(user_db_id)
                test_code = code
                test_title = title
                test_subject = subject or "Noma'lum"
                exam_date = now_str_local()

                tmp_dir = tempfile.gettempdir()
                safe_code = re.sub(r"[^A-Z0-9\-_]+", "_", str(test_code).upper())
                pdf_path = os.path.join(tmp_dir, f"cert_{safe_code}_{user_db_id}.pdf")

                make_certificate_pdf(
                    full_name=user_full_name,
                    subject=test_subject,
                    test_code=test_code,
                    test_title=test_title,
                    exam_date_str=exam_date,
                    out_path=pdf_path
                )

                sent = await message.bot.send_document(
                    chat_id=message.from_user.id,
                    document=FSInputFile(pdf_path),
                    caption="🎁 Sertifikatingiz tayyor! (PDF)"
                )

                cert_file_id = sent.document.file_id if sent and sent.document else None
                async with aiosqlite.connect(DB_NAME) as db2:
                    await db2.execute("""
                        UPDATE results
                        SET certificate_sent=1, certificate_file_id=?, certificate_sent_ts=?
                        WHERE user_id=? AND test_id=?
                    """, (cert_file_id, now_ts(), user_db_id, test_id))
                    await db2.commit()

            except Exception as e:
                # xohlasangiz log qiling
                pass

    await delete_session(user_db_id, test_id, mode)
    cancel_timer(message.from_user.id, test_id)
    await state.clear()

    mistakes_text = "\n".join(mistakes[:25])
    if len(mistakes) > 25:
        mistakes_text += "\n..."

    exam_type = (exam_type or "simple")
    extra_points_line = ""
    if exam_type == "dtm":
        extra_points_line = (
            f"🏁 Ball: <b>{earned_points:.1f}/{max_points:.1f}</b>\n"
            f"📌 Ball foiz: <b>{percent_points:.1f}%</b>\n"
        )

    rasch_pending_block = ""
    if exam_type in ("rasch", "maxsus"):
        rasch_pending_block = (
            "⚡ Rasch foiz: <b>Test tugagach avtomatik hisoblanadi</b>\n"
            "⚡ Rasch ball: <b>Test tugagach avtomatik hisoblanadi</b>\n"
            "🎯 Daraja: <b>Test tugagach chiqadi</b>\n"
            "📌 Rasch ball ko‘rish uchun: <b>“📄 Natijani ko‘rish”</b> tugmasini bosing.\n"
        )


    result_text = (
        f"🏁 <b>Test yakunlandi!</b>\n\n"
        f"🔑 Kod: <code>{code}</code>\n"
        f"📚 Test: <b>{title}</b>\n"
        f"🧮 Savollar: {qcount}\n\n"
        f"✅ To‘g‘ri: <b>{score}/{qcount}</b>\n"
        f"❌ Xato: <b>{qcount - score}</b>\n"
        f"📊 Oddiy foiz: <b>{percent:.1f}%</b>\n"
        f"{extra_points_line}"
        f"{rasch_pending_block}\n"
        f"🔎 <b>Xatolar (qisqa):</b>\n"
        f"{mistakes_text if mistakes_text else 'Xatolar yo‘q ✅'}\n\n"
        f"ℹ️ Bu test bo‘yicha natija faqat <b>1-urinish</b> hisoblanadi."
    )

    # ✅ Natijani ko‘rish tugmasi (test yakunlangach darhol)
    res_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📄 Natijani ko‘rish", callback_data=f"viewres_{test_id}")]
    ])

    admin_flag = await is_admin_user(message.from_user.id)
    await message.answer(
        result_text,
        parse_mode=ParseMode.HTML,
        reply_markup=res_kb
    )
    await message.answer("🏠 Bosh menyu", reply_markup=main_menu_kb(admin_flag))


# ===================== RESULTS MENU =====================
@router.message(F.text == BTN_RESULTS)
async def my_results(message: Message):
    user_db_id = await get_user_db_id(message.from_user.id)
    if not user_db_id:
        return await message.answer("Avval /start orqali ro‘yxatdan o‘ting.")

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT t.id, t.code, t.title, r.score, r.total_questions, r.percent, r.date
            FROM results r
            JOIN tests t ON t.id=r.test_id
            WHERE r.user_id=?
            ORDER BY r.id DESC
            LIMIT 10
        """, (user_db_id,)) as cur:
            rows = await cur.fetchall()

    if not rows:
        return await message.answer("Sizda hali natijalar yo‘q.")

    text = "<b>📄 Natijalar (oxirgi 10)</b>\n\n"
    kb = InlineKeyboardMarkup(inline_keyboard=[])

    for test_id, code, title, score, total, pct, dt in rows:
        text += (
            f"🔹 <code>{code}</code> {title}\n"
            f"   ✅ {score}/{total} | 📊 <b>{float(pct):.1f}%</b> | 🗓 {dt}\n\n"
        )
        kb.inline_keyboard.append([
            InlineKeyboardButton(text=f"📄 Ochish: {code}", callback_data=f"viewres_{test_id}")
        ])

    await message.answer(text, parse_mode=ParseMode.HTML, reply_markup=kb)


# ===================== RATING =====================
@router.message(F.text == BTN_RATING)
async def rating_menu(message: Message):
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🏆 Umumiy reyting (TOP)", callback_data="rt_overall")],
        [InlineKeyboardButton(text="📅 Haftalik TOP (7 kun)", callback_data="rt_weekly")],
        [InlineKeyboardButton(text="📚 Test bo‘yicha reyting", callback_data="rt_tests")]
    ])
    await message.answer("🏆 <b>Reyting</b>\nBo‘limni tanlang:", parse_mode=ParseMode.HTML, reply_markup=kb)


async def _user_names_map():
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT id, full_name FROM users") as cur:
            rows = await cur.fetchall()
    return {r[0]: r[1] for r in rows}


@router.callback_query(F.data == "rt_overall")
async def rating_overall(callback: CallbackQuery):
    user_db_id = await get_user_db_id(callback.from_user.id)
    if not user_db_id:
        return await callback.message.answer("Avval /start orqali ro‘yxatdan o‘ting.")

    names = await _user_names_map()

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT r.user_id, r.percent
            FROM results r
            JOIN (
                SELECT user_id, test_id, MIN(id) AS min_id
                FROM results
                GROUP BY user_id, test_id
            ) x ON x.min_id = r.id
        """) as cur:
            rows = await cur.fetchall()

    if not rows:
        return await callback.message.answer("Hozircha reyting uchun natijalar yo‘q.")

    sum_pct = {}
    cnt = {}
    for uid, pct in rows:
        sum_pct[uid] = sum_pct.get(uid, 0.0) + float(pct)
        cnt[uid] = cnt.get(uid, 0) + 1

    leaderboard = [(uid, sum_pct[uid] / cnt[uid], cnt[uid]) for uid in sum_pct]
    leaderboard.sort(key=lambda x: x[1], reverse=True)

    text = "<b>🏆 Umumiy reyting (TOP)</b>\n(1-urinish natijalari bo‘yicha o‘rtacha %)\n\n"
    for i, (uid, avg, c) in enumerate(leaderboard[:TOP_N], start=1):
        text += f"{i}) {names.get(uid,'Noma’lum')} — <b>{avg:.1f}%</b> (testlar: {c})\n"

    my_rank = None
    my_avg = None
    for i, (uid, avg, c) in enumerate(leaderboard, start=1):
        if uid == user_db_id:
            my_rank, my_avg = i, avg
            break

    if my_rank:
        text += f"\n<b>Sizning o‘rningiz:</b> {my_rank}-o‘rin — {my_avg:.1f}%"
    else:
        text += "\nSiz hali reytingga kirmagansiz."

    await callback.message.answer(text, parse_mode=ParseMode.HTML)


@router.callback_query(F.data == "rt_weekly")
async def rating_weekly(callback: CallbackQuery):
    names = await _user_names_map()
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT r.user_id, r.percent
            FROM results r
            JOIN (
                SELECT user_id, test_id, MIN(id) AS min_id
                FROM results
                GROUP BY user_id, test_id
            ) x ON x.min_id = r.id
            WHERE datetime(r.date) >= datetime('now', '-7 days')
        """) as cur:
            rows = await cur.fetchall()

    if not rows:
        return await callback.message.answer("Oxirgi 7 kunda reyting uchun natija yo‘q.")

    best = {}
    for uid, pct in rows:
        pct = float(pct)
        if uid not in best or pct > best[uid]:
            best[uid] = pct

    leaderboard = sorted([(uid, pct) for uid, pct in best.items()], key=lambda x: x[1], reverse=True)
    text = "<b>📅 Haftalik TOP (7 kun)</b>\n(1-urinish natijalari)\n\n"
    for i, (uid, pct) in enumerate(leaderboard[:TOP_N], start=1):
        text += f"{i}) {names.get(uid,'Noma’lum')} — <b>{pct:.1f}%</b>\n"

    await callback.message.answer(text, parse_mode=ParseMode.HTML)


@router.callback_query(F.data == "rt_tests")
async def rating_tests_list(callback: CallbackQuery):
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT id, code, title FROM tests ORDER BY id DESC LIMIT 30") as cur:
            tests = await cur.fetchall()

    if not tests:
        return await callback.message.answer("Hozircha testlar yo‘q.")

    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for tid, code, title in tests:
        kb.inline_keyboard.append([InlineKeyboardButton(text=f"🔑 {code} — {title}", callback_data=f"rt_test_{tid}")])

    await callback.message.answer("📚 <b>Test bo‘yicha reyting</b>\nTestni tanlang:", parse_mode=ParseMode.HTML, reply_markup=kb)


@router.callback_query(F.data.startswith("rt_test_"))
async def rating_per_test(callback: CallbackQuery):

    test_id = int(callback.data.split("_")[2])

    # ✅ SHU YERGA QO‘SHING
    # ✅ finalize faqat Rasch test uchun
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT exam_type FROM tests WHERE id=?", (test_id,)) as cur:
            row = await cur.fetchone()
        if row:
            et = (row[0] or "simple")
            if et == "rasch":
                await finalize_rush_for_test_if_ready(test_id)
            elif et == "maxsus":
                await finalize_maxsus_for_test_if_ready(test_id)

    names = await _user_names_map()

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT code, title FROM tests WHERE id=?", (test_id,)) as cur:
            t = await cur.fetchone()
        if not t:
            return await callback.message.answer("Test topilmadi.")
        code, title = t

        async with db.execute("""
            SELECT r.user_id, r.percent
            FROM results r
            JOIN (
                SELECT user_id, MIN(id) AS min_id
                FROM results
                WHERE test_id=?
                GROUP BY user_id
            ) x ON x.min_id = r.id
        """, (test_id,)) as cur:
            rows = await cur.fetchall()

    if not rows:
        return await callback.message.answer("Bu test bo‘yicha hali natija yo‘q.")

    leaderboard = sorted([(uid, float(pct)) for uid, pct in rows], key=lambda x: x[1], reverse=True)
    text = f"<b>🏆 Reyting: {title}</b>\n🔑 <code>{code}</code>\n(1-urinish)\n\n"
    for i, (uid, pct) in enumerate(leaderboard[:TOP_N], start=1):
        text += f"{i}) {names.get(uid,'Noma’lum')} — <b>{pct:.1f}%</b>\n"

    await callback.message.answer(text, parse_mode=ParseMode.HTML)


# ===================== ADMIN: PENDING =====================
@router.message(F.text == BTN_ADMIN_PENDING)
async def admin_pending(message: Message):
    if not await is_admin_user(message.from_user.id):
        return

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT p.id, u.full_name, t.code, t.title, t.price, p.created_at
            FROM payments p
            JOIN users u ON u.id=p.user_id
            JOIN tests t ON t.id=p.test_id
            WHERE p.status='pending'
            ORDER BY p.id DESC
            LIMIT 30
        """) as cur:
            rows = await cur.fetchall()

    if not rows:
        return await message.answer("🟡 Pending to‘lovlar yo‘q.")

    text = "<b>🟡 Pending to‘lovlar (oxirgi 30 ta)</b>\n\n"
    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for pid, uname, code, title, price, created_at in rows:
        text += (
            f"🆔 <code>{pid}</code>\n"
            f"👤 {uname}\n"
            f"🔑 <code>{code}</code> — <b>{title}</b>\n"
            f"💰 {price} so'm | 🕒 {created_at}\n\n"
        )
        kb.inline_keyboard.append([InlineKeyboardButton(text=f"Ko‘rish: {pid}", callback_data=f"pview_{pid}")])

    await message.answer(text, parse_mode=ParseMode.HTML, reply_markup=kb)


@router.callback_query(F.data.startswith("pview_"))
async def admin_pending_view(callback: CallbackQuery, bot: Bot):
    if not await is_admin_user(callback.from_user.id):
        return

    pid = int(callback.data.split("_")[1])

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT p.id, p.screenshot_id, p.created_at, u.full_name, t.code, t.title, t.price
            FROM payments p
            JOIN users u ON u.id=p.user_id
            JOIN tests t ON t.id=p.test_id
            WHERE p.id=?
        """, (pid,)) as cur:
            row = await cur.fetchone()

    if not row:
        return await callback.message.answer("Payment topilmadi.")

    pid, screenshot_id, created_at, uname, code, title, price = row
    admin_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Tasdiqlash", callback_data=f"appr_{pid}")],
        [InlineKeyboardButton(text="❌ Rad etish (Sabab)", callback_data=f"rejask_{pid}")]
    ])

    await bot.send_photo(
        chat_id=callback.from_user.id,
        photo=screenshot_id,
        caption=(
            f"🟡 <b>PENDING to‘lov</b>\n\n"
            f"🆔 Payment ID: <code>{pid}</code>\n"
            f"👤 User: {uname}\n"
            f"📚 Test: {title}\n"
            f"🔑 Kod: <code>{code}</code>\n"
            f"💰 Summa: {price} so'm\n"
            f"🕒 Sana: {created_at}"
        ),
        parse_mode=ParseMode.HTML,
        reply_markup=admin_kb
    )


@router.callback_query(F.data.startswith("appr_"))
async def approve_payment(callback: CallbackQuery, bot: Bot):
    if not await is_admin_user(callback.from_user.id):
        return

    payment_id = int(callback.data.split("_")[1])

    async with aiosqlite.connect(DB_NAME) as db:
        await db.execute("UPDATE payments SET status='approved', reject_reason=NULL WHERE id=?", (payment_id,))
        await db.commit()

        async with db.execute("""
            SELECT u.telegram_id, t.code, t.title, t.duration, t.questions_count, t.start_mode, t.start_ts, t.start_at
            FROM payments p
            JOIN users u ON u.id=p.user_id
            JOIN tests t ON t.id=p.test_id
            WHERE p.id=?
        """, (payment_id,)) as cur:
            row = await cur.fetchone()

    if not row:
        return await callback.message.answer("Xatolik: payment topilmadi.")

    user_tg_id, code, title, duration, qcount, start_mode, start_ts_val, start_at = row
    expired, after = is_scheduled_expired(start_mode, start_ts_val, duration)
    if expired:
        await callback.message.answer("❌ Bu test vaqti tugagan. Tasdiqlab bo‘lmaydi.")
        return

    start_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🚀 Testni boshlash", callback_data=f"begin_{payment_id}")]
    ])

    extra = ""
    if (start_mode or "normal") == "scheduled":
        extra = f"\n📅 Boshlanish: <b>{start_at}</b>\n⚠️ PDF faqat vaqt yetganda beriladi."

    await bot.send_message(
        chat_id=user_tg_id,
        text=(
            f"✅ <b>To‘lov tasdiqlandi!</b>\n\n"
            f"📚 Test: <b>{title}</b>\n"
            f"🔑 Kod: <code>{code}</code>\n"
            f"🧮 Savollar: <b>{qcount}</b>\n"
            f"⏳ Vaqt: {duration} daqiqa\n"
            f"{extra}\n\n"
            f"✅ Boshlash uchun tugmani bosing."
        ),
        parse_mode=ParseMode.HTML,
        reply_markup=start_kb
    )

    await callback.message.answer(f"✅ Tasdiqlandi. Payment ID: {payment_id}")


@router.callback_query(F.data.startswith("rejask_"))
async def reject_ask_reason(callback: CallbackQuery, state: FSMContext):
    if not await is_admin_user(callback.from_user.id):
        return
    payment_id = int(callback.data.split("_")[1])
    await state.update_data(payment_id=payment_id)
    await callback.message.answer(
        "❌ <b>Rad etish sababi</b>ni yozing (1-2 jumla):",
        parse_mode=ParseMode.HTML,
        reply_markup=nav_kb(include_back=True)
    )
    await state.set_state(AdminRejectState.waiting_reason)


@router.message(AdminRejectState.waiting_reason)
async def reject_with_reason(message: Message, state: FSMContext, bot: Bot):
    if not await is_admin_user(message.from_user.id):
        await state.clear()
        return
    if (message.text or "") in TOP_MENU_BTNS:
        return

    data = await state.get_data()
    payment_id = int(data.get("payment_id") or 0)
    reason = (message.text or "").strip()
    if not reason:
        return await message.answer("Sabab bo‘sh bo‘lmasin. Qayta yozing:")

    async with aiosqlite.connect(DB_NAME) as db:
        await db.execute("UPDATE payments SET status='rejected', reject_reason=? WHERE id=?", (reason, payment_id))
        await db.commit()

        async with db.execute("""
            SELECT u.telegram_id, t.code, t.title, t.price
            FROM payments p
            JOIN users u ON u.id=p.user_id
            JOIN tests t ON t.id=p.test_id
            WHERE p.id=?
        """, (payment_id,)) as cur:
            row = await cur.fetchone()

    if row:
        user_tg_id, code, title, price = row
        await bot.send_message(
            chat_id=user_tg_id,
            text=(
                f"❌ <b>To‘lov rad etildi</b>\n\n"
                f"📚 Test: <b>{title}</b>\n"
                f"🔑 Kod: <code>{code}</code>\n"
                f"💰 Summa: {price} so'm\n"
                f"📝 Sabab: <i>{reason}</i>\n\n"
                f"📦 'Mening testlarim' bo‘limidan qayta to‘lov yuborishingiz mumkin."
            ),
            parse_mode=ParseMode.HTML
        )

    await message.answer(f"❌ Rad etildi. Payment ID: {payment_id}")
    await state.clear()


# ===================== ADMIN: STATS + EXCEL EXPORT =====================
@router.message(F.text == BTN_ADMIN_STATS)
async def admin_stats(message: Message):
    if not await is_admin_user(message.from_user.id):
        return

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT COUNT(*) FROM users") as cur:
            users_count = (await cur.fetchone())[0]
        async with db.execute("SELECT COUNT(*) FROM tests") as cur:
            tests_count = (await cur.fetchone())[0]
        async with db.execute("SELECT COUNT(*) FROM tests WHERE is_free=1") as cur:
            free_count = (await cur.fetchone())[0]
        async with db.execute("SELECT COUNT(*) FROM tests WHERE is_free=0") as cur:
            paid_count = (await cur.fetchone())[0]

        async with db.execute("SELECT COUNT(*) FROM payments WHERE status='pending'") as cur:
            pending_count = (await cur.fetchone())[0]
        async with db.execute("SELECT COUNT(*) FROM payments WHERE status='approved'") as cur:
            approved_count = (await cur.fetchone())[0]
        async with db.execute("SELECT COUNT(*) FROM payments WHERE status='rejected'") as cur:
            rejected_count = (await cur.fetchone())[0]

        async with db.execute("SELECT percent FROM results") as cur:
            perc = [float(x[0]) for x in await cur.fetchall()]
        overall_avg = (sum(perc) / len(perc)) if perc else 0.0

        async with db.execute("""
            SELECT u.full_name, t.code, t.title, r.percent, r.date
            FROM results r
            JOIN users u ON u.id=r.user_id
            JOIN tests t ON t.id=r.test_id
            ORDER BY r.percent DESC
            LIMIT 1
        """) as cur:
            best = await cur.fetchone()

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📤 Excel eksport (test bo‘yicha)", callback_data="xl_menu")],
        [InlineKeyboardButton(text="⚡ Rasch hisobot (test bo‘yicha)", callback_data="rasch_menu")]
    ])

    text = (
        f"<b>📊 Admin Statistika Paneli</b>\n\n"
        f"👥 Userlar: <b>{users_count}</b>\n"
        f"📚 Testlar: <b>{tests_count}</b> (🆓 {free_count} / 🧾 {paid_count})\n\n"
        f"🟡 Pending: <b>{pending_count}</b>\n"
        f"✅ Approved (sotilgan): <b>{approved_count}</b>\n"
        f"❌ Rejected: <b>{rejected_count}</b>\n\n"
        f"📈 Umumiy o‘rtacha o‘zlashtirish: <b>{overall_avg:.1f}%</b>\n"
    )

    if best:
        uname, code, title, pct, dt = best
        text += f"\n🏅 Eng yuqori natija: <b>{uname}</b>\n🔑 <code>{code}</code> {title} — <b>{float(pct):.1f}%</b> ({dt})\n"

    await message.answer(text, parse_mode=ParseMode.HTML, reply_markup=kb)

@router.callback_query(F.data == "rasch_menu")
async def rasch_menu(callback: CallbackQuery):
    if not await is_admin_user(callback.from_user.id):
        return

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT id, code, title
            FROM tests
            WHERE exam_type='rasch'
            ORDER BY id DESC
            LIMIT 50
        """) as cur:
            tests = await cur.fetchall()

    if not tests:
        return await callback.message.answer("Rasch testlar yo‘q.")

    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for tid, code, title in tests:
        kb.inline_keyboard.append([
            InlineKeyboardButton(text=f"⚡ {code} — {title}", callback_data=f"raschrep_{tid}")
        ])

    await callback.message.answer(
        "⚡ <b>Rasch hisobot</b>\nTestni tanlang:",
        parse_mode=ParseMode.HTML,
        reply_markup=kb
    )

@router.callback_query(F.data == "xl_menu")
async def xl_menu(callback: CallbackQuery):
    if not await is_admin_user(callback.from_user.id):
        return

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT id, code, title FROM tests ORDER BY id DESC LIMIT 50") as cur:
            tests = await cur.fetchall()

    if not tests:
        return await callback.message.answer("Testlar yo‘q.")

    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for tid, code, title in tests:
        kb.inline_keyboard.append([InlineKeyboardButton(text=f"📤 {code} — {title}", callback_data=f"xl_{tid}")])

    await callback.message.answer("📤 <b>Excel eksport</b>\nQaysi test bo‘yicha?", parse_mode=ParseMode.HTML, reply_markup=kb)


@router.callback_query(F.data.startswith("xl_"))
async def xl_export(callback: CallbackQuery, bot: Bot):
    if not await is_admin_user(callback.from_user.id):
        return
    test_id = int(callback.data.split("_")[1])

    # ✅ finalize faqat Rasch test uchun
    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT exam_type FROM tests WHERE id=?", (test_id,)) as cur:
            row = await cur.fetchone()
    if row:
        et = (row[0] or "simple")
        if et == "rasch":
            await finalize_rush_for_test_if_ready(test_id)
        elif et == "maxsus":
            await finalize_maxsus_for_test_if_ready(test_id)

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("""
            SELECT code, title, is_free, price, duration, questions_count, start_mode, start_at
            FROM tests WHERE id=?
        """, (test_id,)) as cur:
            t = await cur.fetchone()

        if not t:
            return await callback.message.answer("Test topilmadi.")

        code, title, is_free, price, duration, qcount, start_mode, start_at = t

        async with db.execute("""
            SELECT u.full_name, u.phone, u.telegram_id,
                   r.score, r.total_questions, r.percent,
                   r.rasch_ball, r.rasch_percent, r.grade,
                   r.date

            FROM results r
            JOIN users u ON u.id=r.user_id
            WHERE r.test_id=?
            ORDER BY r.percent DESC, r.id ASC
        """, (test_id,)) as cur:
            results = await cur.fetchall()

        async with db.execute("""
            SELECT u.full_name, u.phone, u.telegram_id,
                   p.status, p.created_at, p.reject_reason, p.started_at
            FROM payments p
            JOIN users u ON u.id=p.user_id
            WHERE p.test_id=?
            ORDER BY p.id DESC
        """, (test_id,)) as cur:
            pays = await cur.fetchall()

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "TestInfo"

    ws_info.append(["Field", "Value"])
    ws_info.append(["Code", code])
    ws_info.append(["Title", title])
    ws_info.append(["Type", "FREE" if int(is_free) == 1 else "PAID"])
    ws_info.append(["Price", int(price) if price is not None else 0])
    ws_info.append(["Duration(min)", int(duration)])
    ws_info.append(["Questions", int(qcount)])
    # exam_type ni ham yozamiz
    async with aiosqlite.connect(DB_NAME) as db2:
        async with db2.execute("SELECT exam_type FROM tests WHERE id=?", (test_id,)) as cur2:
            etr = await cur2.fetchone()

    ws_info.append(["ExamType", (etr[0] if etr else "simple")])
    ws_info.append(["StartMode", start_mode or "normal"])
    ws_info.append(["StartAt", start_at or ""])

    ws_res = wb.create_sheet("Results")
    ws_res.append(["FullName","Phone","TelegramID","Score","Total","Percent","RaschBall","RaschPercent","Grade","Date"])

    for r in results:
        ws_res.append(list(r))

    ws_pay = wb.create_sheet("Payments")
    ws_pay.append(["FullName", "Phone", "TelegramID", "Status", "CreatedAt", "RejectReason", "StartedAt"])
    for p in pays:
        ws_pay.append(list(p))

    for ws in [ws_info, ws_res, ws_pay]:
        for col in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col)
            for cell in ws[col_letter]:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    safe_code = re.sub(r"[^A-Z0-9\-_]+", "_", code.upper())
    filename = f"test_stats_{safe_code}.xlsx"
    tmp_dir = tempfile.gettempdir()
    path = os.path.join(tmp_dir, filename)
    wb.save(path)

    await bot.send_document(
        chat_id=callback.from_user.id,
        document=FSInputFile(path),
        caption=f"📤 Excel eksport tayyor.\n🔑 <code>{code}</code> — <b>{title}</b>",
        parse_mode=ParseMode.HTML
    )
@router.callback_query(F.data.startswith("raschrep_"))
async def rasch_report_text(callback: CallbackQuery):
    if not await is_admin_user(callback.from_user.id):
        return

    test_id = int(callback.data.split("_")[1])

    # finalize agar vaqti kelgan bo‘lsa
    await finalize_rush_for_test_if_ready(test_id)

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT code, title FROM tests WHERE id=?", (test_id,)) as cur:
            t = await cur.fetchone()

        if not t:
            return await callback.message.answer("Test topilmadi.")

        code, title = t

        async with db.execute("""
        SELECT
            COUNT(*) AS total_results,
            SUM(CASE WHEN rasch_ball IS NOT NULL THEN 1 ELSE 0 END) AS computed,
            AVG(rasch_ball),
            AVG(rasch_percent),
            SUM(CASE WHEN grade IS NOT NULL THEN 1 ELSE 0 END) AS graded
        FROM results
        WHERE test_id=?

        """, (test_id,)) as cur:
            agg = await cur.fetchone()
    if not agg:
        total_results, computed, avg_ball, avg_pct, graded = 0, 0, 0, 0, 0
    else:
        total_results, computed, avg_ball, avg_pct, graded = agg

    text = (
        f"⚡ <b>Rasch Hisobot</b>\n\n"
        f"🔑 <code>{code}</code>\n"
        f"📚 <b>{title}</b>\n\n"
        f"👥 Qatnashganlar: <b>{int(total_results or 0)}</b>\n"
        f"🧮 Hisoblanganlar: <b>{int(computed or 0)}</b>\n"
        f"📈 O‘rtacha Rasch ball: <b>{float(avg_ball or 0):.1f} / 75</b>\n"
        f"📈 O‘rtacha Rasch foiz: <b>{float(avg_pct or 0):.1f}%</b>\n"
        f"🏅 Daraja olganlar: <b>{int(graded or 0)}</b>\n"
    )

    await callback.message.answer(text, parse_mode=ParseMode.HTML)


# ===================== ADMIN: CHANNELS (ADD/EDIT/DELETE/TOGGLE) =====================
@router.message(F.text == BTN_ADMIN_CHANNELS)
async def admin_channels_panel(message: Message, state: FSMContext):
    if not await is_admin_user(message.from_user.id):
        return

    rows = await get_required_channels(active_only=False)
    text = "<b>📣 Majburiy kanallar</b>\n\n"
    if not rows:
        text += "Hozircha kanal qo‘shilmagan.\n"
    else:
        for (_id, chat_id, username, title, join_url, is_active) in rows:
            st = "✅ ON" if int(is_active or 0) == 1 else "⛔ OFF"
            ref = f"@{username}" if username else (str(chat_id) if chat_id else "N/A")
            link = (join_url or "").strip()
            text += f"🆔 <code>{_id}</code> | {st} | <b>{title or ''}</b> | <code>{ref}</code>"
            if link:
                text += f"\n🔗 {link}"
            text += "\n\n"

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ Kanal qo‘shish", callback_data="ch_add")],
        [InlineKeyboardButton(text="🧹 Tahrirlash / O‘chirish", callback_data="ch_manage")],
    ])
    await message.answer(text, parse_mode=ParseMode.HTML, reply_markup=kb)
    await state.clear()


@router.callback_query(F.data == "ch_add")
async def ch_add_start(callback: CallbackQuery, state: FSMContext):
    if not await is_admin_user(callback.from_user.id):
        return
    await callback.message.answer(
        "➕ <b>Kanal/Guruhni kiriting</b>\n\n"
        "✅ Username (public kanal/guruh): <code>@my_channel</code>\n"
        "yoki\n"
        "✅ Chat ID (kanal/guruh): <code>-1001234567890</code>\n\n"
        "⚠️ Eslatma: Guruhda tekshirish ishlashi uchun bot o‘sha guruhga qo‘shilgan bo‘lishi kerak.\n"
        "Kanalda esa bot admin bo‘lishi kerak (a’zolarni ko‘ra olishi uchun).\n\n"
        "Yuboring:",
        parse_mode=ParseMode.HTML,
        reply_markup=nav_kb(include_back=True)
    )
    await state.set_state(AdminChannelState.waiting_channel_ref)


@router.message(AdminChannelState.waiting_channel_ref)
async def ch_add_got_ref(message: Message, state: FSMContext, bot: Bot):
    if not await is_admin_user(message.from_user.id):
        await state.clear()
        return
    if (message.text or "") in TOP_MENU_BTNS:
        return

    ref_raw = (message.text or "").strip()
    ref = _normalize_channel_ref(ref_raw)

    try:
        if _looks_like_int(ref):
            chat = await bot.get_chat(int(ref))
        else:
            chat = await bot.get_chat(f"@{ref}")
    except Exception:
        await message.answer("❌ Kanal topilmadi. Username/ID ni tekshirib qayta yuboring:")
        return

    await state.update_data(
        chat_id=int(chat.id),
        username=(chat.username or None),
        title=(chat.title or "Channel")
    )

    await message.answer(
        "🔗 <b>Join link (ixtiyoriy)</b>\n\n"
        "Public kanal bo‘lsa bo‘sh qoldirsangiz ham bo‘ladi (shunchaki <code>-</code> yuboring).\n"
        "Private kanal bo‘lsa invite link shart.\n\n"
        "Yangi link yuboring yoki <code>-</code> yuboring:",
        parse_mode=ParseMode.HTML,
        reply_markup=nav_kb(include_back=True)
    )
    await state.set_state(AdminChannelState.waiting_join_url)


@router.message(AdminChannelState.waiting_join_url)
async def ch_add_save(message: Message, state: FSMContext):
    if not await is_admin_user(message.from_user.id):
        await state.clear()
        return
    if (message.text or "") in TOP_MENU_BTNS:
        return

    join_url = (message.text or "").strip()
    if join_url == "-":
        join_url = ""

    if join_url and not join_url.startswith("http"):
        await message.answer("⚠️ Link noto‘g‘ri. Masalan: https://t.me/my_channel yoki invite link. Qayta yuboring:")
        return

    data = await state.get_data()
    chat_id = int(data["chat_id"])
    username = data.get("username")
    title = data.get("title")

    async with aiosqlite.connect(DB_NAME) as db:
        await db.execute("""
            INSERT INTO required_channels (chat_id, username, title, join_url, is_active)
            VALUES (?, ?, ?, ?, 1)
            ON CONFLICT(chat_id) DO UPDATE SET
                username=excluded.username,
                title=excluded.title,
                join_url=excluded.join_url,
                is_active=1
        """, (chat_id, username, title, join_url or None))
        await db.commit()

    await state.clear()
    await message.answer(f"✅ Qo‘shildi: <b>{title}</b>", parse_mode=ParseMode.HTML, reply_markup=main_menu_kb(True))


@router.callback_query(F.data == "ch_manage")
async def ch_manage(callback: CallbackQuery):
    if not await is_admin_user(callback.from_user.id):
        return

    rows = await get_required_channels(active_only=False)
    if not rows:
        return await callback.message.answer("Ro‘yxat bo‘sh.")

    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for (_id, chat_id, username, title, join_url, is_active) in rows:
        st = "ON✅" if int(is_active or 0) == 1 else "OFF⛔"
        name = title or (f"@{username}" if username else str(chat_id))
        kb.inline_keyboard.append([
            InlineKeyboardButton(text=f"{st} {name}", callback_data=f"ch_toggle_{_id}"),
            InlineKeyboardButton(text="✏️", callback_data=f"ch_edit_{_id}"),
            InlineKeyboardButton(text="🗑", callback_data=f"ch_del_{_id}")
        ])

    await callback.message.answer("🧹 <b>Tahrirlash</b>\nON/OFF, ✏️ tahrirlash yoki 🗑 o‘chirish:", parse_mode=ParseMode.HTML, reply_markup=kb)


@router.callback_query(F.data.startswith("ch_toggle_"))
async def ch_toggle(callback: CallbackQuery):
    if not await is_admin_user(callback.from_user.id):
        return
    _id = int(callback.data.split("_")[2])

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT is_active FROM required_channels WHERE id=?", (_id,)) as cur:
            row = await cur.fetchone()
        if not row:
            return await callback.answer("Topilmadi", show_alert=True)

        new_val = 0 if int(row[0] or 0) == 1 else 1
        await db.execute("UPDATE required_channels SET is_active=? WHERE id=?", (new_val, _id))
        await db.commit()

    await callback.answer("✅ Yangilandi", show_alert=False)


@router.callback_query(F.data.startswith("ch_del_"))
async def ch_delete(callback: CallbackQuery):
    if not await is_admin_user(callback.from_user.id):
        return
    _id = int(callback.data.split("_")[2])

    async with aiosqlite.connect(DB_NAME) as db:
        await db.execute("DELETE FROM required_channels WHERE id=?", (_id,))
        await db.commit()

    await callback.answer("🗑 O‘chirildi", show_alert=False)


@router.callback_query(F.data.startswith("ch_edit_"))
async def ch_edit_start(callback: CallbackQuery, state: FSMContext):
    if not await is_admin_user(callback.from_user.id):
        return
    _id = int(callback.data.split("_")[2])
    await state.update_data(edit_id=_id)
    await callback.message.answer(
        "✏️ <b>Tahrirlash</b>\n\n"
        "2 qatorda yuboring:\n"
        "1) Yangi title (o‘zgartirmasangiz <code>-</code>)\n"
        "2) Yangi join link:\n"
        "   • o‘zgartirmasangiz <code>-</code>\n"
        "   • tozalash (o‘chirish) uchun <code>clear</code>\n\n"
        "Misol (faqat linkni tozalash):\n"
        "<code>-</code>\n"
        "<code>clear</code>\n\n"
        "Misol (title+link yangilash):\n"
        "<code>DTM Kanal</code>\n"
        "<code>https://t.me/dtm_demo</code>",
        parse_mode=ParseMode.HTML,
        reply_markup=nav_kb(include_back=True)
    )
    await state.set_state(AdminChannelState.waiting_edit_payload)

@router.message(AdminChannelState.waiting_edit_payload)
async def ch_edit_save(message: Message, state: FSMContext):
    if not await is_admin_user(message.from_user.id):
        await state.clear()
        return
    if (message.text or "") in TOP_MENU_BTNS:
        return

    lines = [x.strip() for x in (message.text or "").splitlines() if x.strip()]
    if len(lines) < 2:
        await message.answer("⚠️ 2 qatorda yuboring: (title) va (link). Qayta yuboring:")
        return

    new_title_raw = lines[0]
    new_link_raw = lines[1]

    # 1) TITLE: "-" => o'zgarmasin
    new_title = None if new_title_raw == "-" else new_title_raw

    # 2) LINK:
    # "-"     => o'zgarmasin
    # "clear" => tozalansin (bo'sh/NULL)
    # else    => yangi link (https://... bo'lishi shart)
    keep_link = False
    clear_link = False
    new_link = None

    if new_link_raw == "-":
        keep_link = True
    elif new_link_raw.lower() == "clear":
        clear_link = True
    else:
        if not new_link_raw.startswith("http"):
            await message.answer("⚠️ Link noto‘g‘ri. https://... bo‘lishi kerak. Qayta yuboring:")
            return
        new_link = new_link_raw

    data = await state.get_data()
    _id = int(data.get("edit_id") or 0)

    async with aiosqlite.connect(DB_NAME) as db:
        async with db.execute("SELECT title, join_url FROM required_channels WHERE id=?", (_id,)) as cur:
            row = await cur.fetchone()
        if not row:
            await state.clear()
            await message.answer("Topilmadi.")
            return

        old_title, old_link = row

        final_title = new_title if new_title is not None else old_title

        if keep_link:
            final_link = old_link
        elif clear_link:
            final_link = ""  # xohlasangiz None qilsangiz DBda NULL bo'ladi
        else:
            final_link = new_link

        await db.execute(
            "UPDATE required_channels SET title=?, join_url=? WHERE id=?",
            (final_title, final_link, _id)
        )
        await db.commit()

    await state.clear()
    await message.answer("✅ Tahrirlandi.", reply_markup=main_menu_kb(True))




DTM_DEFAULT_CFG = {
    "subjects": [
        {"name": "1-fan", "q": 30, "w": 3.1},
        {"name": "2-fan", "q": 30, "w": 2.1},
        {"name": "3-fan", "q": 10, "w": 1.1},
        {"name": "4-fan", "q": 10, "w": 1.1},
        {"name": "5-fan", "q": 10, "w": 1.1},
    ]
}

def dtm_total_questions(cfg: dict) -> int:
    return sum(int(s["q"]) for s in cfg["subjects"])
def dtm_score_points(user_ans: str, key: str, cfg_json: str) -> tuple[float, float, dict]:
    cfg = json.loads(cfg_json) if cfg_json else DTM_DEFAULT_CFG
    idx = 0
    earned = 0.0
    maxp = 0.0
    details = {"subjects": []}

    for s in cfg["subjects"]:
        q = int(s["q"])
        w = float(s["w"])
        part_user = user_ans[idx:idx+q]
        part_key = key[idx:idx+q]

        correct = 0
        for i in range(q):
            if part_user[i] == part_key[i]:
                correct += 1

        earned_s = correct * w
        max_s = q * w

        earned += earned_s
        maxp += max_s

        details["subjects"].append({
            "name": s.get("name"),
            "q": q,
            "w": w,
            "correct": correct,
            "earned": earned_s,
            "max": max_s
        })
        idx += q

    return earned, maxp, details


# ===================== ADMIN: ADD TEST (DYNAMIC QCOUNT + SCHEDULE) =====================
@router.message(F.text == BTN_ADMIN_ADD)
async def admin_add_test(message: Message, state: FSMContext):
    if not await is_admin_user(message.from_user.id):
        return
    await message.answer(
        "📚 <b>Test qaysi fan bo‘yicha?</b>\n"
        "Masalan: Matematika, Ingliz tili, Tarix, Ona tili...\n\n"
        "Fan nomini kiriting:",
        parse_mode=ParseMode.HTML,
        reply_markup=nav_kb(include_back=True)
    )
    await state.set_state(AdminAddTestState.subject)

    @router.message(AdminAddTestState.subject)
    async def set_subject(message: Message, state: FSMContext):
        if (message.text or "") in TOP_MENU_BTNS:
            return
        subject = (message.text or "").strip()
        if len(subject) < 2:
            return await message.answer("⚠️ Fan nomi juda qisqa. Qayta kiriting:")
        await state.update_data(subject=subject)

        await message.answer(
            "✅ Endi test nomini kiriting (masalan: DTM Blok 2024):",
            reply_markup=nav_kb(include_back=True)
        )
        await state.set_state(AdminAddTestState.title)


@router.message(AdminAddTestState.title)
async def set_title(message: Message, state: FSMContext):
    if (message.text or "") in TOP_MENU_BTNS:
        return
    await state.update_data(title=(message.text or "").strip())
    await message.answer(
        "🔑 <b>Test uchun UNIQUE KOD kiriting</b>\n"
        "Masalan: <code>DTM24A</code>, <code>MS-001</code>, <code>FREE45</code>\n\n"
        "Qoidalar: 3-20 belgi, faqat A-Z, 0-9, -, _",
        parse_mode=ParseMode.HTML,
        reply_markup=nav_kb(include_back=True)
    )
    await state.set_state(AdminAddTestState.code)


@router.message(AdminAddTestState.code)
async def set_code(message: Message, state: FSMContext):
    if (message.text or "") in TOP_MENU_BTNS:
        return
    code = normalize_code(message.text)
    if not is_valid_code(code):
        await message.answer("⚠️ Kod formati noto‘g‘ri. Qayta kiriting:")
        return
    if await test_code_exists(code):
        await message.answer("⚠️ Bu kod mavjud. Boshqa kod kiriting:")
        return

    await state.update_data(code=code)
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🆓 Tekin test", callback_data="setfree_1")],
        [InlineKeyboardButton(text="🧾 Pullik test", callback_data="setfree_0")]
    ])
    await message.answer("Test turi (to‘lov bo‘yicha):", reply_markup=kb)
    await state.set_state(AdminAddTestState.is_free)


@router.callback_query(F.data.startswith("setfree_"))
async def set_is_free(callback: CallbackQuery, state: FSMContext):
    is_free = int(callback.data.split("_")[1])
    await state.update_data(is_free=is_free)

    if is_free == 1:
        await callback.message.answer("✅ Tekin test.\n⏳ Vaqt limitini kiriting (daqiqada):", reply_markup=nav_kb(include_back=True))
        await state.set_state(AdminAddTestState.duration)
    else:
        await callback.message.answer("✅ Pullik test.\n💰 Narxni kiriting (so'm):", reply_markup=nav_kb(include_back=True))
        await state.set_state(AdminAddTestState.price)


@router.message(AdminAddTestState.price)
async def set_price(message: Message, state: FSMContext):
    if (message.text or "") in TOP_MENU_BTNS:
        return
    if not (message.text or "").isdigit():
        return await message.answer("Faqat raqam kiriting.")
    await state.update_data(price=int(message.text))
    await message.answer("⏳ Vaqt limitini kiriting (daqiqada):", reply_markup=nav_kb(include_back=True))
    await state.set_state(AdminAddTestState.duration)


@router.message(AdminAddTestState.duration)
async def set_duration(message: Message, state: FSMContext):
    if (message.text or "") in TOP_MENU_BTNS:
        return
    if not (message.text or "").isdigit():
        return await message.answer("Faqat raqam kiriting.")
    await state.update_data(duration=int(message.text))

    await message.answer(
        "🧮 Savollar sonini kiriting (masalan 30, 45, 60):",
        reply_markup=nav_kb(include_back=True)
    )
    await state.set_state(AdminAddTestState.questions_count)


@router.message(AdminAddTestState.questions_count)
async def set_qcount(message: Message, state: FSMContext):
    if (message.text or "") in TOP_MENU_BTNS:
        return
    if not (message.text or "").isdigit():
        return await message.answer("Faqat raqam kiriting.")
    qcount = int(message.text)
    if qcount < 1 or qcount > 300:
        return await message.answer("⚠️ Savollar soni 1..300 oralig‘ida bo‘lsin.")

    await state.update_data(questions_count=qcount)

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📝 Oddiy test", callback_data="etype_simple")],
        [InlineKeyboardButton(text="⚡ Rasch (MS)", callback_data="etype_rasch")],
        [InlineKeyboardButton(text="⭐ Maxsus (Rasch-like)", callback_data="etype_maxsus")],
        [InlineKeyboardButton(text="📚 DTM (5 fan)", callback_data="etype_dtm")],
    ])
    await message.answer("Test rejimini tanlang:", reply_markup=kb)
    await state.set_state(AdminAddTestState.exam_type)

@router.callback_query(F.data.startswith("etype_"))
async def set_exam_type(callback: CallbackQuery, state: FSMContext):
    et = callback.data.split("_", 1)[1]  # simple/rasch/dtm

    if et == "dtm":
        cfg = DTM_DEFAULT_CFG
        await state.update_data(exam_type="dtm", dtm_cfg=json.dumps(cfg, ensure_ascii=False))
        # ✅ DTMda savollar soni majburiy 90
        await state.update_data(questions_count=dtm_total_questions(cfg))

        await callback.message.answer(
            "✅ DTM (5 fan) rejimi tanlandi.\n"
            "Majburiy: 1-2 fan 30 tadan, 3-4-5 fan 10 tadan.\n"
            "Ball: 3.1 / 2.1 / 1.1 / 1.1 / 1.1\n"
            f"Jami savol: {dtm_total_questions(cfg)}"
        )

    elif et == "rasch":
        await state.update_data(exam_type="rasch", dtm_cfg=None)
        await callback.message.answer("✅ Rasch (MS) rejimi tanlandi.")
    elif et == "maxsus":
        await state.update_data(exam_type="maxsus", dtm_cfg=None)
        await callback.message.answer("✅ Maxsus rejim tanlandi. Natija test tugagach Rasch ko‘rinishida chiqariladi.")
    else:
        await state.update_data(exam_type="simple", dtm_cfg=None)
        await callback.message.answer("✅ Oddiy rejim tanlandi.")

    # Endi schedule tanlatamiz (eski logika)
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Oddiy test (darhol)", callback_data="sch_normal")],
        [InlineKeyboardButton(text="📅 Rejalashtirilgan test", callback_data="sch_scheduled")]
    ])
    await callback.message.answer("Test boshlanish rejimi:", reply_markup=kb)
    await state.set_state(AdminAddTestState.schedule_mode)

@router.callback_query(F.data == "sch_normal")
async def schedule_normal(callback: CallbackQuery, state: FSMContext):
    await state.update_data(start_mode="normal", start_ts=0, start_at=None)
    await callback.message.answer("📄 Test PDF faylini yuklang:", reply_markup=nav_kb(include_back=True))
    await state.set_state(AdminAddTestState.file)


@router.callback_query(F.data == "sch_scheduled")
async def schedule_scheduled(callback: CallbackQuery, state: FSMContext):
    await state.update_data(start_mode="scheduled")
    await callback.message.answer(
        "📅 Boshlanish sanasini kiriting (YYYY-MM-DD)\nMasalan: <code>2026-02-15</code>",
        parse_mode=ParseMode.HTML,
        reply_markup=nav_kb(include_back=True)
    )
    await state.set_state(AdminAddTestState.schedule_date)


@router.message(AdminAddTestState.schedule_date)
async def set_schedule_date(message: Message, state: FSMContext):
    if (message.text or "") in TOP_MENU_BTNS:
        return
    if not parse_date(message.text or ""):
        return await message.answer("⚠️ Sana formati noto‘g‘ri. Masalan: 2026-02-15")
    await state.update_data(schedule_date=message.text.strip())
    await message.answer(
        "🕒 Boshlanish vaqtini kiriting (HH:MM)\nMasalan: <code>16:00</code>",
        parse_mode=ParseMode.HTML,
        reply_markup=nav_kb(include_back=True)
    )
    await state.set_state(AdminAddTestState.schedule_time)


@router.message(AdminAddTestState.schedule_time)
async def set_schedule_time(message: Message, state: FSMContext):
    if (message.text or "") in TOP_MENU_BTNS:
        return
    if not parse_time(message.text or ""):
        return await message.answer("⚠️ Vaqt formati noto‘g‘ri. Masalan: 16:00")
    data = await state.get_data()
    res = to_start_ts(data.get("schedule_date", ""), (message.text or "").strip())
    if not res:
        return await message.answer("⚠️ Sana/Vaqt xato. Qayta kiriting.")

    start_ts_val, start_at = res
    await state.update_data(start_ts=start_ts_val, start_at=start_at)

    await message.answer(
        f"✅ Rejalashtirildi.\n📅 Boshlanish: <b>{start_at}</b>\n\n📄 Endi PDF yuklang:",
        parse_mode=ParseMode.HTML,
        reply_markup=nav_kb(include_back=True)
    )
    await state.set_state(AdminAddTestState.file)


@router.message(AdminAddTestState.file, F.document)
async def set_file(message: Message, state: FSMContext):
    await state.update_data(file_id=message.document.file_id)
    data = await state.get_data()
    qcount = int(data.get("questions_count") or 30)

    await message.answer(
        f"✅ Endi javob kalitini kiriting.\n"
        f"⚠️ Javoblar soni aynan <b>{qcount}</b> ta bo‘lishi kerak.\n"
        f"Masalan: <code>abcdabcd...</code>",
        parse_mode=ParseMode.HTML,
        reply_markup=nav_kb(include_back=True)
    )
    await state.set_state(AdminAddTestState.answers)


@router.message(AdminAddTestState.answers)
async def save_test(message: Message, state: FSMContext):
    if not await is_admin_user(message.from_user.id):
        await state.clear()
        return
    if (message.text or "") in TOP_MENU_BTNS:
        return

    data = await state.get_data()
    answers = normalize_answers(message.text)
    qcount = int(data.get("questions_count") or 30)

    if not is_answer_string_valid(answers):
        await message.answer("⚠️ Javob kaliti faqat harflardan iborat bo‘lsin. Qayta yuboring:")
        return

    if len(answers) != qcount:
        await message.answer(
            f"⚠️ Javoblar soni noto‘g‘ri.\n"
            f"Kerakli: <b>{qcount}</b>, Siz yubordingiz: <b>{len(answers)}</b>.\n\n"
            f"Iltimos, aniq <b>{qcount} ta</b> javob yuboring:",
            parse_mode=ParseMode.HTML
        )
        return

    is_free = int(data.get("is_free") or 0)
    price = 0 if is_free else int(data.get("price") or 0)

    start_mode = data.get("start_mode") or "normal"
    start_ts_val = int(data.get("start_ts") or 0)
    start_at = data.get("start_at")

    async with aiosqlite.connect(DB_NAME) as db:
        exam_type = data.get("exam_type") or "simple"
        dtm_cfg = data.get("dtm_cfg")

        await db.execute("""
            INSERT INTO tests (
                code, title, subject, price, duration, questions_count, file_id, answers, is_free,
                start_mode, start_ts, start_at,
                exam_type, dtm_cfg
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            data.get("code"),
            data.get("title"),
            data.get("subject"),  # ✅ NEW
            price,
            int(data.get("duration") or 0),
            qcount,
            data.get("file_id"),
            answers,
            is_free,
            start_mode,
            start_ts_val,
            start_at,
            exam_type,
            dtm_cfg
        ))

        await db.commit()

    extra = ""
    if start_mode == "scheduled":
        extra = f"\n📅 Boshlanish: <b>{start_at}</b>"

    await message.answer(
        f"✅ Test qo‘shildi!\n🔑 Kod: <code>{data.get('code')}</code>\n🧮 Savollar: <b>{qcount}</b>{extra}",
        parse_mode=ParseMode.HTML,
        reply_markup=main_menu_kb(True)
    )
    await state.clear()


# ===================== HELP =====================
@router.message(F.text == BTN_HELP)
async def help_menu(message: Message):
    await message.answer(
        "ℹ️ <b>Yordam</b>\n\n"
        "🔎 Kod orqali test topish:\n"
        "• 🧾 Pullik testlar → 'Kod orqali topish'\n"
        "• 🆓 Tekin testlar → 'Kod orqali topish'\n"
        "• 📦 Mening testlarim → 'Kod orqali topish'\n\n"
        "✅ Har bir test bo‘yicha natija faqat <b>1-urinish</b> qabul qilinadi.\n"
        "✅ Savollar soni nechta bo‘lsa, javob ham aynan shuncha bo‘lishi shart.\n\n"
        "📅 Rejalashtirilgan test bo‘lsa, PDF faqat boshlanish vaqti kelganda beriladi.\n\n"
        "🔐 Xavfsizlik: spam bo‘lsa vaqtincha blok beradi.\n"
        "👤Admin: @jasur_aktamov\n"
        "📞Admin: +998932244730\n"
        f"🔁 Navigatsiya: {BTN_HOME} / {BTN_BACK}",
        parse_mode=ParseMode.HTML
    )

CERT_CHECK_EVERY_SEC = 30

async def certificate_daemon(bot: Bot):
    while True:
        try:
            async with aiosqlite.connect(DB_NAME) as db:
                # 1) Sertifikat yuborilmagan natijalarni topamiz
                async with db.execute("""
                    SELECT r.user_id, r.test_id, r.percent, r.grade, r.date,
                           u.telegram_id, u.full_name,
                           t.code, t.title, t.subject, t.exam_type
                    FROM results r
                    JOIN users u ON u.id=r.user_id
                    JOIN tests t ON t.id=r.test_id
                    WHERE (r.certificate_sent IS NULL OR r.certificate_sent=0)
                    ORDER BY r.id ASC
                    LIMIT 20
                """) as cur:
                    rows = await cur.fetchall()

            for (user_id_db, test_id, pct, grade, date_str,
                 user_tg_id, full_name,
                 tcode, ttitle, tsubject, exam_type) in rows:

                et = (exam_type or "simple")

                # 2) Rasch/Maxsus bo'lsa finalize bo'lishi shart (test tugagan bo'lsa)
                if et == "rasch":
                    await finalize_rush_for_test_if_ready(test_id)
                    # finalize bo'lgach grade qayta o'qiladi
                    async with aiosqlite.connect(DB_NAME) as db2:
                        async with db2.execute("""
                            SELECT rasch_percent, grade
                            FROM results
                            WHERE user_id=? AND test_id=?
                            ORDER BY id ASC LIMIT 1
                        """, (user_id_db, test_id)) as cur2:
                            rr = await cur2.fetchone()
                    if rr:
                        pct = rr[0]
                        grade = rr[1]

                elif et == "maxsus":
                    await finalize_maxsus_for_test_if_ready(test_id)
                    async with aiosqlite.connect(DB_NAME) as db2:
                        async with db2.execute("""
                            SELECT rasch_percent, grade
                            FROM results
                            WHERE user_id=? AND test_id=?
                            ORDER BY id ASC LIMIT 1
                        """, (user_id_db, test_id)) as cur2:
                            rr = await cur2.fetchone()
                    if rr:
                        pct = rr[0]
                        grade = rr[1]

                # 3) Eligibility
                if not is_certificate_eligible(et, float(pct or 0.0), grade):
                    # sertifikat yo'q -> yuborilgan deb belgilamaymiz (xohlasangiz belgilab qo'yish mumkin)
                    continue

                # 4) PDF yasab yuboramiz
                tmp_dir = tempfile.gettempdir()
                safe_code = re.sub(r"[^A-Z0-9\-_]+", "_", str(tcode).upper())
                pdf_path = os.path.join(tmp_dir, f"cert_{safe_code}_{user_id_db}.pdf")

                make_certificate_pdf(
                    full_name=full_name or "Noma'lum",
                    subject=(tsubject or "Noma'lum"),
                    test_code=tcode,
                    test_title=ttitle,
                    exam_date_str=(date_str or now_str_local()),
                    out_path=pdf_path
                )

                sent = await bot.send_document(
                    chat_id=int(user_tg_id),
                    document=FSInputFile(pdf_path),
                    caption="🎁 Sertifikatingiz tayyor! (PDF)"
                )
                cert_file_id = sent.document.file_id if sent and sent.document else None

                async with aiosqlite.connect(DB_NAME) as db3:
                    await db3.execute("""
                        UPDATE results
                        SET certificate_sent=1, certificate_file_id=?, certificate_sent_ts=?
                        WHERE user_id=? AND test_id=?
                    """, (cert_file_id, now_ts(), user_id_db, test_id))
                    await db3.commit()

        except Exception:
            pass

        await asyncio.sleep(CERT_CHECK_EVERY_SEC)
# ===================== MAIN =====================
async def main():
    await init_db()
    bot = Bot(token=BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
    asyncio.create_task(certificate_daemon(bot))
    dp = Dispatcher()
    dp.include_router(router)

    logging.basicConfig(level=logging.INFO, stream=sys.stdout)

    await dp.start_polling(bot)


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("Bot to'xtatildi")
